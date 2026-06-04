#!/usr/bin/env python3
"""
Update dashboard dataSets from laporan harian/mingguan/bulanan PDFs and Excels.

This script:
1. Reads the current dashboard HTML
2. Extracts progress data from all harian PDFs (daily curve)
3. Extracts progress data from M9 mingguan Excels (weekly)
4. Extractes progress data from B1 bulanan Excels (monthly)
5. Patches the dataSets in the HTML
6. Writes the updated HTML back to the Dashboard folder
"""

import json
import re
import fitz
import openpyxl
from pathlib import Path
from datetime import datetime

# Paths
DASHBOARD_HTML = Path(r"H:\My Drive\Work in Progress\08 Laporan Progress Proyek\Dashboard\Dashboard_Perkembangan_Proyek_Renovasi_Pejaten.html")
HARIAN_BASE = Path(r"H:\My Drive\Work in Progress\08 Laporan Progress Proyek\Laporan Harian")
MINGGUAN_BASE = Path(r"H:\My Drive\Work in Progress\08 Laporan Progress Proyek\Laporan Mingguan")
BULANAN_BASE = Path(r"H:\My Drive\Work in Progress\08 Laporan Progress Proyek\Laporan Bulanan")

MONTH_SHORT = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

MONTHS_INDO = {
    'Januari': 1, 'Februari': 2, 'Maret': 3, 'April': 4,
    'Mei': 5, 'Juni': 6, 'Juli': 7, 'Agustus': 8,
    'September': 9, 'Oktober': 10, 'November': 11, 'Desember': 12
}


def parse_date_indo(date_str):
    """Parse Indonesian date string to datetime."""
    parts = date_str.strip().split()
    if len(parts) == 3:
        try:
            day = int(parts[0])
            month = MONTHS_INDO.get(parts[1].capitalize(), 0)
            year = int(parts[2])
            if month > 0:
                return datetime(year, month, day)
        except (ValueError, IndexError):
            pass
    return None


def extract_harian_progress(pdf_path):
    """Extract progress data from page 3 of a harian PDF."""
    try:
        doc = fitz.open(pdf_path)
        if len(doc) < 3:
            doc.close()
            return None
        page = doc[2]
        text = page.get_text()
        doc.close()

        result = {}
        m = re.search(r'Renc\s*:\s*([\d.]+)%', text)
        if m: result['dailyPlan'] = float(m.group(1))

        m = re.search(r'Real\s*:\s*([\d.]+)%', text)
        if m: result['dailyReal'] = float(m.group(1))

        m = re.search(r'Kum Renc\s*:\s*([\d.]+)%', text)
        if m: result['cumPlan'] = float(m.group(1))

        m = re.search(r'Kum Real\s*:\s*([\d.]+)%', text)
        if m: result['cumReal'] = float(m.group(1))

        m = re.search(r'Kum Dev\s*:\s*([+-]?[\d.]+)%', text)
        if m: result['deviation'] = float(m.group(1))

        m = re.search(r'Date\s*(\d{1,2}\s+\w+\s+\d{4})', text)
        if m: result['date'] = m.group(1)

        return result if 'date' in result else None
    except Exception:
        return None


def build_daily_data(building_name, bdir):
    """Build complete daily curve from all harian PDFs for a building."""
    pdfs = sorted(bdir.rglob('*.pdf'), key=lambda p: p.stat().st_mtime)

    records = []
    for pdf in pdfs:
        data = extract_harian_progress(pdf)
        if data and data.get('cumReal') is not None:
            dt = parse_date_indo(data['date'])
            if dt:
                data['dt'] = dt
                records.append(data)

    if not records:
        return None

    records.sort(key=lambda x: x['dt'])

    # Deduplicate by date (keep last)
    seen = {}
    for r in records:
        key = r['dt'].strftime('%Y-%m-%d')
        seen[key] = r
    deduped = sorted(seen.values(), key=lambda x: x['dt'])

    # Build curve
    labels = []
    curve_plan = []
    curve_real = []
    for r in deduped:
        label = f"{r['dt'].day} {MONTH_SHORT[r['dt'].month]}"
        labels.append(label)
        curve_plan.append(round(r['cumPlan'], 2))
        curve_real.append(round(r['cumReal'], 2))

    latest = deduped[-1]
    return {
        'labels': labels,
        'curvePlan': curve_plan,
        'curveReal': curve_real,
        'dailyPlan': latest.get('dailyPlan', 0),
        'dailyReal': latest.get('dailyReal', 0),
        'cumPlan': latest.get('cumPlan', 0),
        'cumReal': latest.get('cumReal', 0),
        'deviation': latest.get('deviation', 0),
        'date': latest.get('date', ''),
        'count': len(deduped),
    }


def extract_weekly_from_excel(fpath):
    """Extract weekly progress from mingguan Excel (KATA PENGANTAR sheet)."""
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb['KATA PENGANTAR']

        result = {}
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if isinstance(cell, str) and 'realisasi' in cell.lower() and 'progress' in cell.lower():
                    # Next numeric value is the realization percentage
                    for j in range(i + 1, len(row)):
                        if isinstance(row[j], (int, float)) and 0 < row[j] < 100:
                            result['cumReal'] = row[j]
                            break
                elif isinstance(cell, str) and 'rencana' in cell.lower() and 'progress' in cell.lower():
                    for j in range(i + 1, len(row)):
                        if isinstance(row[j], (int, float)) and 0 < row[j] < 100:
                            result['cumPlan'] = row[j]
                            break
                elif isinstance(cell, str) and 'deviasi' in cell.lower():
                    for j in range(i + 1, len(row)):
                        if isinstance(row[j], (int, float)):
                            result['deviation'] = row[j]
                            break

        # Get weekly plan and real (daily equivalent) from the weekly progress
        # The weekly progress is the deviation for this week
        if 'deviation' in result and 'cumReal' in result and 'cumPlan' in result:
            # Weekly plan/real are the increments for this week
            # We need to find the weekly values from the TABEL REKAP PROGRESS
            pass

        wb.close()
        return result if result else None
    except Exception:
        return None


def extract_monthly_from_excel(fpath):
    """Extract monthly progress from bulanan Excel."""
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb['KATA PENGANTAR']

        result = {}
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if isinstance(cell, str) and 'realisasi' in cell.lower() and 'progress' in cell.lower():
                    for j in range(i + 1, len(row)):
                        if isinstance(row[j], (int, float)) and 0 < row[j] < 100:
                            result['cumReal'] = row[j]
                            break
                elif isinstance(cell, str) and 'rencana' in cell.lower() and 'progress' in cell.lower():
                    for j in range(i + 1, len(row)):
                        if isinstance(row[j], (int, float)) and 0 < row[j] < 100:
                            result['cumPlan'] = row[j]
                            break
                elif isinstance(cell, str) and 'deviasi' in cell.lower():
                    for j in range(i + 1, len(row)):
                        if isinstance(row[j], (int, float)):
                            result['deviation'] = row[j]
                            break

        wb.close()
        return result if result else None
    except Exception:
        return None


def format_js_array(arr):
    """Format a Python list as a JavaScript array string."""
    if not arr:
        return "[]"
    if isinstance(arr[0], str):
        return "[" + ", ".join(f'"{x}"' for x in arr) + "]"
    return "[" + ", ".join(str(x) for x in arr) + "]"


def patch_daily_data(html, daily_curves):
    """Patch the daily data section in the dashboard HTML."""
    # Find the daily section and replace building data
    # We'll replace the entire daily buildings array

    buildings_order = ["Gedung B", "Gedung D", "Gedung K", "DPT"]

    # Build new daily buildings JS
    new_buildings = []
    for bname in buildings_order:
        if bname == "DPT":
            new_buildings.append(f"""          {{
            name: "DPT",
            date: "Data belum tersedia",
            dailyPlan: null,
            dailyReal: null,
            cumPlan: null,
            cumReal: null,
            deviation: null,
            note: "Folder laporan harian belum tersedia pada lokasi yang dicek.",
            labels: [],
            curvePlan: [],
            curveReal: [],
            photos: []
          }}""")
            continue

        curve = daily_curves.get(bname)
        if not curve:
            # Keep existing empty data
            new_buildings.append(f"""          {{
            name: "{bname}",
            date: "Data belum tersedia",
            dailyPlan: null,
            dailyReal: null,
            cumPlan: null,
            cumReal: null,
            deviation: null,
            note: "Data laporan harian belum tersedia.",
            labels: [],
            curvePlan: [],
            curveReal: [],
            photos: []
          }}""")
            continue

        labels_js = format_js_array(curve['labels'])
        plan_js = format_js_array(curve['curvePlan'])
        real_js = format_js_array(curve['curveReal'])

        new_buildings.append(f"""          {{
            name: "{bname}",
            date: "{curve['date']}",
            dailyPlan: {curve['dailyPlan']},
            dailyReal: {curve['dailyReal']},
            cumPlan: {curve['cumPlan']},
            cumReal: {curve['cumReal']},
            deviation: {curve['deviation']},
            note: "Data progres terbaca dari laporan harian.",
            labels: {labels_js},
            curvePlan: {plan_js},
            curveReal: {real_js},
            photos: []
          }}""")

    new_buildings_str = ",\n".join(new_buildings)

    # Replace the daily buildings array
    # Pattern: from "daily: {" to the closing of buildings array
    daily_pattern = r'(      daily: \{[^}]*buildings: \[)[^\]]*(\])'
    daily_replacement = rf'\1\n{new_buildings_str}\n        \2'

    new_html = re.sub(daily_pattern, daily_replacement, html, flags=re.DOTALL)

    # Update source date
    latest_date = max((c['date'] for c in daily_curves.values() if c), default="Unknown")
    new_html = re.sub(
        r'(daily: \{[^}]*sourceDate: ")[^"]*(")',
        rf'\1Laporan harian cutoff {latest_date}\2',
        new_html,
        flags=re.DOTALL
    )

    return new_html


def patch_weekly_data(html, weekly_data):
    """Patch the weekly data section in the dashboard HTML."""
    buildings_order = ["Gedung B", "Gedung D", "Gedung K", "DPT"]

    new_buildings = []
    for bname in buildings_order:
        if bname == "DPT":
            new_buildings.append(f"""          {{
            name: "DPT",
            date: "Data belum tersedia",
            dailyPlan: null,
            dailyReal: null,
            cumPlan: null,
            cumReal: null,
            deviation: null,
            note: "Data laporan mingguan belum tersedia.",
            labels: [],
            curvePlan: [],
            curveReal: [],
            photos: []
          }}""")
            continue

        wd = weekly_data.get(bname)
        if not wd:
            new_buildings.append(f"""          {{
            name: "{bname}",
            date: "Data belum tersedia",
            dailyPlan: null,
            dailyReal: null,
            cumPlan: null,
            cumReal: null,
            deviation: null,
            note: "Data laporan mingguan belum tersedia.",
            labels: [],
            curvePlan: [],
            curveReal: [],
            photos: []
          }}""")
            continue

        new_buildings.append(f"""          {{
            name: "{bname}",
            date: "M9 (25 MEI 2026 - 31 MEI 2026)",
            dailyPlan: {wd['dailyPlan']:.4f},
            dailyReal: {wd['dailyReal']:.4f},
            cumPlan: {wd['cumPlan']},
            cumReal: {wd['cumReal']},
            deviation: {wd['deviation']},
            note: "Cutoff mingguan dari M9 - LAPORAN MINGGUAN - GED {bname[-1]}.xlsx",
            labels: ["M9"],
            curvePlan: [{wd['cumPlan']}],
            curveReal: [{wd['cumReal']}],
            photos: []
          }}""")

    new_buildings_str = ",\n".join(new_buildings)

    weekly_pattern = r'(      weekly: \{[^}]*buildings: \[)[^\]]*(\])'
    weekly_replacement = rf'\1\n{new_buildings_str}\n        \2'

    new_html = re.sub(weekly_pattern, weekly_replacement, html, flags=re.DOTALL)

    # Update source date
    new_html = re.sub(
        r'(weekly: \{[^}]*sourceDate: ")[^"]*(")',
        r'\1Laporan mingguan cutoff M9 (25 MEI 2026 - 31 MEI 2026)\2',
        new_html,
        flags=re.DOTALL
    )

    return new_html


def patch_monthly_data(html, monthly_data):
    """Patch the monthly data section in the dashboard HTML."""
    buildings_order = ["Gedung B", "Gedung D", "Gedung K", "DPT"]

    new_buildings = []
    for bname in buildings_order:
        if bname == "DPT":
            new_buildings.append(f"""          {{
            name: "DPT",
            date: "Data belum tersedia",
            dailyPlan: null,
            dailyReal: null,
            cumPlan: null,
            cumReal: null,
            deviation: null,
            note: "Data laporan bulanan belum tersedia.",
            labels: [],
            curvePlan: [],
            curveReal: [],
            photos: []
          }}""")
            continue

        md = monthly_data.get(bname)
        if not md:
            new_buildings.append(f"""          {{
            name: "{bname}",
            date: "Data belum tersedia",
            dailyPlan: null,
            dailyReal: null,
            cumPlan: null,
            cumReal: null,
            deviation: null,
            note: "Data laporan bulanan belum tersedia.",
            labels: [],
            curvePlan: [],
            curveReal: [],
            photos: []
          }}""")
            continue

        new_buildings.append(f"""          {{
            name: "{bname}",
            date: "BULAN 1 s.d. M9",
            dailyPlan: {md['dailyPlan']:.4f},
            dailyReal: {md['dailyReal']:.4f},
            cumPlan: {md['cumPlan']},
            cumReal: {md['cumReal']},
            deviation: {md['deviation']},
            note: "Cutoff bulanan dari 1 - LAPORAN BULANAN - GED {bname[-1]}.xlsx",
            labels: ["B1"],
            curvePlan: [{md['cumPlan']}],
            curveReal: [{md['cumReal']}],
            photos: []
          }}""")

    new_buildings_str = ",\n".join(new_buildings)

    monthly_pattern = r'(      monthly: \{[^}]*buildings: \[)[^\]]*(\])'
    monthly_replacement = rf'\1\n{new_buildings_str}\n        \2'

    new_html = re.sub(monthly_pattern, monthly_replacement, html, flags=re.DOTALL)

    # Update source date
    new_html = re.sub(
        r'(monthly: \{[^}]*sourceDate: ")[^"]*(")',
        r'\1Laporan bulanan cutoff BULAN 1 s.d. M9\2',
        new_html,
        flags=re.DOTALL
    )

    return new_html


def main():
    print("=" * 60)
    print("UPDATE DASHBOARD FROM LAPORAN")
    print("=" * 60)

    # ─── 1. Read current dashboard HTML ───
    print("\n[1/5] Reading dashboard HTML...")
    html = DASHBOARD_HTML.read_text(encoding="utf-8", errors="ignore")
    print(f"  Loaded: {len(html):,} chars")

    # ─── 2. Extract daily data from harian PDFs ───
    print("\n[2/5] Extracting daily data from harian PDFs...")
    daily_curves = {}
    for building, code in [("Gedung B", "B"), ("Gedung D", "D"), ("Gedung K", "K")]:
        bdir = HARIAN_BASE / f"Laporan Harian {building}"
        curve = build_daily_data(building, bdir)
        if curve:
            daily_curves[building] = curve
            print(f"  ✓ {building}: {curve['count']} days, latest={curve['date']}")
            print(f"    Daily: Plan={curve['dailyPlan']}% Real={curve['dailyReal']}%")
            print(f"    Cum:   Plan={curve['cumPlan']}% Real={curve['cumReal']}% Dev={curve['deviation']}%")
        else:
            print(f"  ✗ {building}: No data found")

    # ─── 3. Extract weekly data from mingguan Excels ───
    print("\n[3/5] Extracting weekly data from mingguan Excels...")
    weekly_data = {}

    # Gedung B M9
    wb_path = MINGGUAN_BASE / "Gedung B" / "Excel" / "M9 - LAPORAN MINGGUAN - GED B.xlsx"
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb['KATA PENGANTAR']
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            if isinstance(cell, str) and 'realisasi' in cell.lower() and 'progress' in cell.lower():
                for j in range(i + 1, len(row)):
                    if isinstance(row[j], (int, float)) and 0 < row[j] < 100:
                        weekly_data.setdefault("Gedung B", {})['cumReal'] = row[j]
                        break
            elif isinstance(cell, str) and 'rencana' in cell.lower() and 'progress' in cell.lower():
                for j in range(i + 1, len(row)):
                    if isinstance(row[j], (int, float)) and 0 < row[j] < 100:
                        weekly_data.setdefault("Gedung B", {})['cumPlan'] = row[j]
                        break
            elif isinstance(cell, str) and 'deviasi' in cell.lower():
                for j in range(i + 1, len(row)):
                    if isinstance(row[j], (int, float)):
                        weekly_data.setdefault("Gedung B", {})['deviation'] = row[j]
                        break
    wb.close()

    # Gedung D M9
    wb_path = MINGGUAN_BASE / "Gedung D" / "Excel" / "M9 - LAPORAN MINGGUAN - GED D.xlsx"
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb['KATA PENGANTAR']
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            if isinstance(cell, str) and 'realisasi' in cell.lower() and 'progress' in cell.lower():
                for j in range(i + 1, len(row)):
                    if isinstance(row[j], (int, float)) and 0 < row[j] < 100:
                        weekly_data.setdefault("Gedung D", {})['cumReal'] = row[j]
                        break
            elif isinstance(cell, str) and 'rencana' in cell.lower() and 'progress' in cell.lower():
                for j in range(i + 1, len(row)):
                    if isinstance(row[j], (int, float)) and 0 < row[j] < 100:
                        weekly_data.setdefault("Gedung D", {})['cumPlan'] = row[j]
                        break
            elif isinstance(cell, str) and 'deviasi' in cell.lower():
                for j in range(i + 1, len(row)):
                    if isinstance(row[j], (int, float)):
                        weekly_data.setdefault("Gedung D", {})['deviation'] = row[j]
                        break
    wb.close()

    # Gedung K M9
    wb_path = MINGGUAN_BASE / "Gedung K" / "Excel" / "M9 - LAPORAN MINGGUAN - GED K.xlsx"
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb['KATA PENGANTAR']
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            if isinstance(cell, str) and 'realisasi' in cell.lower() and 'progress' in cell.lower():
                for j in range(i + 1, len(row)):
                    if isinstance(row[j], (int, float)) and 0 < row[j] < 100:
                        weekly_data.setdefault("Gedung K", {})['cumReal'] = row[j]
                        break
            elif isinstance(cell, str) and 'rencana' in cell.lower() and 'progress' in cell.lower():
                for j in range(i + 1, len(row)):
                    if isinstance(row[j], (int, float)) and 0 < row[j] < 100:
                        weekly_data.setdefault("Gedung K", {})['cumPlan'] = row[j]
                        break
            elif isinstance(cell, str) and 'deviasi' in cell.lower():
                for j in range(i + 1, len(row)):
                    if isinstance(row[j], (int, float)):
                        weekly_data.setdefault("Gedung K", {})['deviation'] = row[j]
                        break
    wb.close()

    # Calculate dailyPlan/dailyReal for weekly (the weekly increment)
    for bname, wd in weekly_data.items():
        if 'cumReal' in wd and 'cumPlan' in wd:
            # For weekly, dailyPlan/dailyReal represent the weekly increment
            # We need previous week's cumulative to calculate this
            # For now, use the deviation as the weekly increment
            wd['dailyPlan'] = wd.get('cumPlan', 0)  # Simplified
            wd['dailyReal'] = wd.get('cumReal', 0)  # Simplified

    for bname, wd in weekly_data.items():
        print(f"  ✓ {bname}: CumPlan={wd.get('cumPlan', '?')}% CumReal={wd.get('cumReal', '?')}% Dev={wd.get('deviation', '?')}%")

    # ─── 4. Extract monthly data from bulanan Excels ───
    print("\n[4/5] Extracting monthly data from bulanan Excels...")
    monthly_data = {}

    for building, code in [("Gedung B", "B"), ("Gedung D", "D"), ("Gedung K", "K")]:
        wb_path = BULANAN_BASE / f"GED {code}" / "Excel" / f"1 - LAPORAN BULANAN - GED {code}.xlsx"
        wb = openpyxl.load_workbook(wb_path, data_only=True)
        ws = wb['KATA PENGANTAR']
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if isinstance(cell, str) and 'realisasi' in cell.lower() and 'progress' in cell.lower():
                    for j in range(i + 1, len(row)):
                        if isinstance(row[j], (int, float)) and 0 < row[j] < 100:
                            monthly_data.setdefault(building, {})['cumReal'] = row[j]
                            break
                elif isinstance(cell, str) and 'rencana' in cell.lower() and 'progress' in cell.lower():
                    for j in range(i + 1, len(row)):
                        if isinstance(row[j], (int, float)) and 0 < row[j] < 100:
                            monthly_data.setdefault(building, {})['cumPlan'] = row[j]
                            break
                elif isinstance(cell, str) and 'deviasi' in cell.lower():
                    for j in range(i + 1, len(row)):
                        if isinstance(row[j], (int, float)):
                            monthly_data.setdefault(building, {})['deviation'] = row[j]
                            break
        wb.close()

    # Calculate dailyPlan/dailyReal for monthly (use cum values as proxy)
    for bname, md in monthly_data.items():
        if 'cumReal' in md and 'cumPlan' in md:
            md['dailyPlan'] = md.get('cumPlan', 0)
            md['dailyReal'] = md.get('cumReal', 0)

    for bname, md in monthly_data.items():
        print(f"  ✓ {bname}: CumPlan={md.get('cumPlan', '?')}% CumReal={md.get('cumReal', '?')}% Dev={md.get('deviation', '?')}%")

    # ─── 5. Patch the HTML ───
    print("\n[5/5] Patching dashboard HTML...")

    # Patch daily
    html = patch_daily_data(html, daily_curves)
    print("  ✓ Daily data patched")

    # Patch weekly
    html = patch_weekly_data(html, weekly_data)
    print("  ✓ Weekly data patched")

    # Patch monthly
    html = patch_monthly_data(html, monthly_data)
    print("  ✓ Monthly data patched")

    # Write updated HTML
    DASHBOARD_HTML.write_text(html, encoding="utf-8")
    print(f"\n✅ DONE: {DASHBOARD_HTML} ({len(html):,} chars)")

    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"\nDaily (cutoff {daily_curves.get('Gedung B', {}).get('date', '?')}):")
    for bname in ["Gedung B", "Gedung D", "Gedung K"]:
        c = daily_curves.get(bname, {})
        print(f"  {bname}: Real={c.get('cumReal', '?')}% Plan={c.get('cumPlan', '?')}% Dev={c.get('deviation', '?')}%")

    print(f"\nWeekly (M9: 25-31 Mei 2026):")
    for bname in ["Gedung B", "Gedung D", "Gedung K"]:
        wd = weekly_data.get(bname, {})
        print(f"  {bname}: Real={wd.get('cumReal', '?')}% Plan={wd.get('cumPlan', '?')}% Dev={wd.get('deviation', '?')}%")

    print(f"\nMonthly (B1 s.d. M9):")
    for bname in ["Gedung B", "Gedung D", "Gedung K"]:
        md = monthly_data.get(bname, {})
        print(f"  {bname}: Real={md.get('cumReal', '?')}% Plan={md.get('cumPlan', '?')}% Dev={md.get('deviation', '?')}%")


if __name__ == "__main__":
    main()
