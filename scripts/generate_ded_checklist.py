#!/usr/bin/env python3
"""Generate DED checklist Excel from folder scan."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True, size=10)
sub_fill = PatternFill("solid", fgColor="2E75B6")
sub_font = Font(color="FFFFFF", bold=True, size=9)
green_fill = PatternFill("solid", fgColor="C6EFCE")
green_font = Font(color="006100", size=9)
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
yellow_font = Font(color="9C5700", size=9)
red_fill = PatternFill("solid", fgColor="FFC7CE")
red_font = Font(color="9C0006", size=9)
na_fill = PatternFill("solid", fgColor="D9D9D9")
na_font = Font(color="7F7F7F", size=9, italic=True)
bold9 = Font(bold=True, size=9)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, val, fill=header_fill, font=header_font):
    cell.value = val
    cell.fill = fill
    cell.font = font
    cell.alignment = center
    cell.border = border

def dat(cell, val, fill=None, font=None, align=left):
    cell.value = val
    if fill: cell.fill = fill
    if font: cell.font = font
    cell.alignment = align
    cell.border = border

def yn(val):
    """Return (display, fill, font) for Yes/No/Partial."""
    if val == "✅ ADA":
        return val, green_fill, green_font
    elif val == "📤 SUBMITTED":
        return val, yellow_fill, yellow_font
    elif val == "❌ TIDAK ADA":
        return val, red_fill, red_font
    else:
        return val, na_fill, na_font

# ═══════════════════════════════════════════════
# SHEET 1 — SUMMARY
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = "SUMMARY"
ws.sheet_properties.tabColor = "1F4E79"

# Column widths
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 14
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 14
ws.column_dimensions["I"].width = 14

# Row heights
ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 18

# Title
ws.merge_cells("A1:I1")
t = ws["A1"]
t.value = "DED CHECKLIST — PROJECT PEJATEN (BIN RENOVASI)"
t.font = Font(bold=True, size=14, color="1F4E79")
t.alignment = center

ws.merge_cells("A2:I2")
s = ws["A2"]
s.value = "Lokasi: H:\\My Drive\\Work in Progress\\02 DED (gambar dari Perencana)"
s.font = Font(size=9, color="666666", italic=True)
s.alignment = center

# Header row
row = 3
ws.row_dimensions[row].height = 40
headers = ["DISIPLIN / DOKUMEN", "GEDUNG B", "GEDUNG D", "GEDUNG K", "INTERIOR (B/D/K)", "RKS", "RAB", "LAPORAN", "OUTLINE SPEK"]
for col, h in enumerate(headers, 1):
    hdr(ws.cell(row, col), h)

# Data — highlevel per gedung based on folder scan
# Key: (discipline, gedung_b_status, gedung_d_status, gedung_k_status, interior, rks, rab, laporan, outline_spek)
data = [
    # ── ARSITEKTUR ──
    ("🏛️ ARSITEKTUR", "", "", "", "", "", "", "", ""),
    ("  - Denah / Site Plan", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Tampak & Potongan", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Pola Lantai", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Plafond", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Kusen / Pintu", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Toilet Detail", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Tangga Detail", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Fasad Detail", "", "✅ ADA", "", "", "", "", "", ""),
    ("  - Schedule Material", "✅ ADA", "", "", "", "", "", "", ""),
    ("  - Ars PDF consolidated", "✅ ADA", "", "", "", "", "", "", ""),
    ("  - Ars DWG count", "~15 DWG", "~25 DWG", "~20 DWG", "", "", "", "", ""),
    # ── STRUKTUR ──
    ("🔧 STRUKTUR", "", "", "", "", "", "", "", ""),
    ("  - Pondasi", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Sloof / Balok", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Kolom", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Plat Lantai", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Portal / Rangka Atap", "✅ ADA", "", "✅ ADA", "", "", "", "", ""),
    ("  - Standar Detail (SD-01~06)", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Struktur PDF", "✅ ADA", "", "✅ ADA", "", "", "", "", ""),
    ("  - Struktur DWG count", "~5 DWG", "", "~3 DWG", "", "", "", "", ""),
    # ── MEP ──
    ("⚡ MEP (Mekanikal Elektronik)", "", "", "", "", "", "", "", ""),
    ("  - Plumbing (PL)", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Pemadam Kebakaran (PK)", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - AC / VAC", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Elektrikal (LAK)", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Elektronik / Low Voltage", "✅ ADA", "✅ ADA", "✅ ADA", "", "", "", "", ""),
    ("  - Lift / TDG", "", "", "✅ ADA", "", "", "", "", ""),
    ("  - MEP PDF", "✅ ADA", "", "✅ ADA", "", "", "", "", ""),
    ("  - MEP DWG count", "~15 DWG", "", "~40 DWG", "", "", "", "", ""),
    # ── INTERIOR ──
    ("🏠 INTERIOR", "", "", "", "", "", "", "", ""),
    ("  - Denah Interior", "", "", "", "✅ ADA (D, K)", "", "", "", ""),
    ("  - Detail Ruangan", "", "", "", "✅ ADA (D)", "", "", "", ""),
    ("  - Built-in Furniture", "", "", "", "✅ ADA (K)", "", "", "", ""),
    ("  - Loose Furniture", "", "", "", "✅ ADA (K)", "", "", "", ""),
    ("  - BOQ / RKS Interior", "", "", "", "✅ ADA (D)", "", "", "", ""),
    # ── SITE DEV ──
    ("🏗️ SITE DEVELOPMENT", "", "", "", "", "", "", "", ""),
    ("  - Site Plan / Grading", "✅ ADA", "✅ ADA", "", "", "", "", "", ""),
    ("  - PDF", "✅ ADA", "✅ ADA", "", "", "", "", "", ""),
    # ── RKS ──
    ("📋 RKS", "", "", "", "", "", "", "", ""),
    ("  - RKS Arsitektur", "", "", "", "", "✅ ADA (B)", "", "", ""),
    ("  - RKS Interior", "", "", "", "", "✅ ADA (D)", "", "", ""),
    ("  - RKS MEP Mekanikal", "", "", "", "", "", "", "", ""),
    ("  - RKS Elektrikal", "", "", "", "", "✅ ADA (D)", "", "", ""),
    # ── RAB ──
    ("💰 RAB", "", "", "", "", "", "", "", ""),
    ("  - RAB Pekerjaan Fisik", "", "", "", "", "", "✅ ADA (B,K)", "", ""),
    ("  - RAB Interior", "", "", "", "", "", "✅ ADA (K)", "", ""),
    ("  - BOQ MEP", "", "", "", "", "", "✅ ADA (D)", "", ""),
    # ── LAPORAN ──
    ("📝 LAPORAN", "", "", "", "", "", "", "", ""),
    ("  - Laporan Pendahuluan", "", "", "", "", "", "", "✅ ADA (B)", ""),
    ("  - Laporan Antara", "", "", "", "", "", "", "✅ ADA (B)", ""),
    ("  - Laporan Akhir (Draft)", "", "", "", "", "", "", "✅ ADA (B)", ""),
    ("  - Laporan Akhir (Final)", "", "", "", "", "", "", "✅ ADA (B)", ""),
    # ── OUTLINE SPEK ──
    ("📄 OUTLINE SPESIFIKASI", "", "", "", "", "", "", "", ""),
    ("  - Outline Spek Gedung", "✅ ADA", "", "✅ ADA", "", "", "", "", ""),
]

for i, row_data in enumerate(data, row + 1):
    ws.row_dimensions[i].height = 20
    for j, val in enumerate(row_data, 1):
        cell = ws.cell(i, j)
        if j == 1:
            # Discipline column
            if val and val[0] in "🏛🔧⚡🏠🏗📋💰📝📄":
                cell.value = val
                cell.fill = sub_fill
                cell.font = sub_font
                cell.alignment = left
            elif val.startswith("  -"):
                cell.value = val
                cell.alignment = left
                cell.font = Font(size=9)
            else:
                cell.value = val
                cell.font = Font(size=9)
        else:
            # Status columns
            if val == "✅ ADA":
                dat(cell, val, green_fill, green_font, center)
            elif val == "❌ TIDAK ADA":
                dat(cell, "—", na_fill, na_font, center)
            elif "DWG" in val or "PDF" in val:
                dat(cell, val, PatternFill("solid", fgColor="DEEAF1"), Font(size=8, color="2E75B6"), center)
            elif val == "📤 SUBMITTED":
                dat(cell, val, yellow_fill, yellow_font, center)
            elif val and val != "":
                dat(cell, val, yellow_fill, yellow_font, center)
            else:
                dat(cell, "—", na_fill, na_font, center)

# Freeze panes
ws.freeze_panes = "B4"

# ═══════════════════════════════════════════════
# SHEET 2 — DETAIL GEDUNG B
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("GEDUNG B")
ws2.sheet_properties.tabColor = "2E75B6"
ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 45
ws2.column_dimensions["D"].width = 20
ws2.column_dimensions["E"].width = 12

ws2.merge_cells("A1:E1")
ws2["A1"].value = "GEDUNG B — Detail File DED"
ws2["A1"].font = Font(bold=True, size=12, color="1F4E79")

h2 = ["NO", "KATEGORI", "NAMA FILE / FOLDER", "TIPE", "KET"]
for col, h in enumerate(h2, 1):
    hdr(ws2.cell(2, col), h)

b_data = [
    ("1", "ARSITEKTUR", "GAMBAR DED/ARSITEKTUR gedung B/", "Folder", "15+ DWG + PDF"),
    ("", "", "01. DENAH BONGKARAN EXISTING.dwg", "DWG", ""),
    ("", "", "02. SITEPLAN & BLOCK PLAN GEDUNG B.dwg", "DWG", ""),
    ("", "", "03. DENAH GEDUNG B.dwg", "DWG", ""),
    ("", "", "A-03. TAMPAK.dwg", "DWG", ""),
    ("", "", "A-05. RENC. POLA LANTAI.dwg", "DWG", ""),
    ("", "", "A-06. RENC. PLAFOND.dwg", "DWG", ""),
    ("", "", "A-07. SCHEDULE MATERIAL GD B.dwg", "DWG", ""),
    ("", "", "A-10. DETAIL TOILET GEDUNG B-02.dwg", "DWG", ""),
    ("", "", "Detail Kusen Gedung B/", "Folder", "DWG + PDF"),
    ("", "", "DENAH KUSEN Gedung B/", "Folder", ""),
    ("", "", "ARS POT GEDUNG.B.25.03.26.dwg", "DWG", ""),
    ("", "", "ARS TANGGA GEDUNG.B.25.03.26 FINAL REVISI.dwg", "DWG", ""),
    ("", "", "ARS GEDUNG B.pdf", "PDF", "Konsolidated"),
    ("2", "STRUKTUR", "GAMBAR DED/STRUKTUR/", "Folder", "DWG + PDF"),
    ("", "", "STR- GEDUNG B.dwg", "DWG", "Struktur utama"),
    ("", "", "STRUKTUR - GEDUNG B update.pdf", "PDF", ""),
    ("", "", "REVISI STR PONDASI GEDUNG B 150526/", "Folder", "Revisi May 2026"),
    ("", "", "REVISI STR PONDASI GEDUNG B 150526.dwg", "DWG", ""),
    ("", "", "DED GEDUNG B/STRUKTUR - GEDUNG B update.pdf", "PDF", "Copy terbaru"),
    ("3", "MEP", "GAMBAR DED/MEP/", "Folder", "5 DWG + PDF"),
    ("", "", "1.INSTALASI PLAMBING GEDUNG B.dwg", "DWG", ""),
    ("", "", "2.INSTALASI PEMADAM KEBAKARAN GEDUNG B.dwg", "DWG", ""),
    ("", "", "3.INSTALASI AIR CONDITIONING & VENTILASI.dwg", "DWG", ""),
    ("", "", "4.INSTALASI ELEKTRIKAL GEDUNG B.dwg", "DWG", "Rev 1"),
    ("", "", "5.INSTALASI ELEKTRONIK GDUNG B.dwg", "DWG", ""),
    ("", "", "MEP GEDUNG B.pdf", "PDF", "Konsolidated"),
    ("4", "SITE DEV", "GAMBAR DED/SITE DEVELOPMENT/", "Folder", "DWG + PDF"),
    ("", "", "01. SITE DEVELOPMENT.dwg", "DWG", ""),
    ("", "", "Site Development GEDUNG B.pdf", "PDF", ""),
    ("5", "DED (folder)", "DED GEDUNG B/", "Folder", "Kumpulan DED terbaru"),
    ("", "", "20260429 DED INTERIOR OFFICE.pdf", "PDF", "Interior Office"),
    ("", "", "Site Development GEDUNG B.pdf", "PDF", ""),
    ("", "", "STRUKTUR - GEDUNG B update.pdf", "PDF", ""),
    ("6", "RKS", "RKS/RKS GEDUNG B (DEPUTI-I).pdf", "PDF", ""),
    ("7", "RAB", "RAB/RAB_GEDUNG B_7.pdf", "PDF", ""),
    ("", "", "RAB/RAB_GEDUNG B_7.xlsx", "XLSX", ""),
    ("8", "LAPORAN", "LAPORAN/", "Folder", "4 dokumen"),
    ("", "", "1. Laporan Pendahuluan/Laporan Pendahuluan.pdf", "PDF", ""),
    ("", "", "2. Laporan Antara/Laporan Antara.pdf", "PDF", ""),
    ("", "", "3. Draft Laporan Akhir/Draft Laporan Akhir.pdf", "PDF", ""),
    ("", "", "4. Laporan Akhir/Laporan Akhir.pdf", "PDF", ""),
    ("9", "OUTLINE SPEK", "Outline Spesifikasi Gedung B R.1.pdf", "PDF", ""),
    ("10", "LAINNYA", "R.GWT DAN STP 6022026/", "Folder", "GWT + Pompa"),
    ("", "", "Laporan Pengukuran Lokasi BIN - Gedung B - Pejaten.pdf", "PDF", ""),
    ("", "", "COVER.pdf (GAMBAR DED)", "PDF", "Cover gambar"),
]

for i, row_data in enumerate(b_data, 3):
    ws2.row_dimensions[i].height = 18
    for j, val in enumerate(row_data, 1):
        cell = ws2.cell(i, j)
        if j == 2 and val:
            cell.value = val
            cell.fill = sub_fill
            cell.font = sub_font
            cell.alignment = left
        elif val.endswith("/"):
            cell.value = val
            cell.font = Font(bold=True, size=9, color="2E75B6")
            cell.alignment = left
        else:
            cell.value = val
            cell.font = Font(size=9)
            cell.alignment = left if j != 1 else center
        cell.border = border

ws2.freeze_panes = "A3"

# ═══════════════════════════════════════════════
# SHEET 3 — DETAIL GEDUNG D
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet("GEDUNG D")
ws3.sheet_properties.tabColor = "375623"
ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 45
ws3.column_dimensions["D"].width = 20
ws3.column_dimensions["E"].width = 12

ws3.merge_cells("A1:E1")
ws3["A1"].value = "GEDUNG D — Detail File DED"
ws3["A1"].font = Font(bold=True, size=12, color="375623")

for col, h in enumerate(h2, 1):
    c = ws3.cell(2, col)
    c.value = h
    c.fill = PatternFill("solid", fgColor="375623")
    c.font = Font(color="FFFFFF", bold=True, size=10)
    c.alignment = center
    c.border = border

d_data = [
    ("1", "ARSITEKTUR", "01. ARSITEKTUR gedung D/", "Folder", "25+ DWG"),
    ("", "", "01. SITE PLAN & BLOK PLAN/", "Folder", "DWG"),
    ("", "", "02. DENAH (INC. PARSIAL)/", "Folder", "DWG"),
    ("", "", "03. TAMPAK & POTONGAN/", "Folder", "DWG"),
    ("", "", "04. PARSIAL FASAD/", "Folder", "DWG"),
    ("", "", "05. DETAIL FASAD/", "Folder", "DWG"),
    ("", "", "06. DETAIL CORE/", "Folder", "DWG"),
    ("", "", "07. DETAIL TANGGA/", "Folder", "DWG"),
    ("", "", "08. POLA LANTAI (INC. PARSIAL)/", "Folder", "DWG"),
    ("", "", "09. POLA PLAFON (INC. PARSIAL)/", "Folder", "DWG"),
    ("", "", "10. SKEMA KUSEN/", "Folder", "DWG"),
    ("", "", "  - Ars RKS & Outline Spek", "Folder", "RKS"),
    ("2", "STRUKTUR", "03. STRUKTUR/", "Folder", "CAD + PDF"),
    ("", "", "(folder kosong di scan — cek subfolder)", "", ""),
    ("3", "INTERIOR", "02. INTERIOR/", "Folder", "CAD + PDF"),
    ("", "", "INT 0 - INFORMASI UMUM/", "Folder", "Cover, Scope, Drawing List"),
    ("", "", "INT 1 - DENAH GENERAL/", "Folder", "Denah interior"),
    ("", "", "INT 2 - DETAIL RUANGAN/", "Folder", "14 ruangan detail"),
    ("", "", "  - Detail: LOBBY, DIREKTUR, DEPUTI, STAFF, RAPAT, ESELON, dll", "", ""),
    ("", "", "03. RKS INTERIOR BIN - GEDUNG D.pdf", "PDF", "RKS Interior"),
    ("", "", "BQ Interior gedung D.xlsx", "XLSX", ""),
    ("4", "MEP MEKANIKAL", "05. MEKANIKAL/", "Folder", "CAD + PDF"),
    ("", "", "01. AC/", "Folder", "DWG"),
    ("", "", "02. LIFT/", "Folder", "DWG"),
    ("", "", "03. PLUMBING/", "Folder", "DWG"),
    ("", "", "04. HYDRANT/", "Folder", "DWG"),
    ("", "", "05. SPRINKLER/", "Folder", "DWG"),
    ("", "", "DED MEKANIKAL GEDUNG KANTOR D.pdf", "PDF", ""),
    ("", "", "RKS & OUTLINE SPEK MEKANIKAL.pdf", "PDF", ""),
    ("5", "MEP ELEKTRIKAL", "06. ELEKTRIKAL/", "Folder", "CAD + PDF"),
    ("", "", "LAK (Elektrikal Daya)/", "Folder", "9 subfolder"),
    ("", "", "  - SITEPLAN, SYSTEM, PENCAHAYAAN, KOTAK KONTAK", "", ""),
    ("", "", "  - SCHEDULE PANEL, R.GENSET, PEREDAM PETIR, KABEL TRAY, DETAIL", "", ""),
    ("", "", "LAL (Low Voltage)/", "Folder", "5 subfolder"),
    ("", "", "  - FIRE ALARM, DATA/TELPON/IPTV, SECURITY, CONFERENCE, DAFTAR GAMBAR", "", ""),
    ("", "", "DED ELEKTRIKAL GEDUNG KANTOR D.pdf", "PDF", ""),
    ("", "", "RKS & OUTLINE SPEK ELEKTRIKAL.pdf", "PDF", ""),
    ("6", "INFRASTRUKTUR", "04. INFRASTRUKTUR/", "Folder", "CAD + PDF"),
    ("", "", "RKS_Infrastruktur Gedung D.pdf", "PDF", ""),
    ("", "", "PDF GRADING GEDUNG D.pdf", "PDF", ""),
    ("7", "SITE DEV", "PDF GRADING GEDUNG D.pdf", "PDF", "Site Grading"),
    ("8", "RAB / BOQ", "2026.05.22 RAB GD. KANTOR BIN (Konstruksi).xlsx", "XLSX", ""),
    ("", "", "2026.05.22 RAB GD. KANTOR BIN (Pengadaan).xlsx", "XLSX", ""),
    ("", "", "BQ Interior gedung D.xlsx", "XLSX", ""),
    ("9", "LAINNYA", "STP Biofilter.dwg", "DWG", ""),
    ("", "", "2026.05.11_Gedung D_Revisi GWT.dwg", "DWG", "Revisi GWT"),
]

for i, row_data in enumerate(d_data, 3):
    ws3.row_dimensions[i].height = 18
    for j, val in enumerate(row_data, 1):
        cell = ws3.cell(i, j)
        if j == 2 and val:
            cell.value = val
            cell.fill = PatternFill("solid", fgColor="548235")
            cell.font = Font(color="FFFFFF", bold=True, size=9)
            cell.alignment = left
        elif val.endswith("/"):
            cell.value = val
            cell.font = Font(bold=True, size=9, color="375623")
            cell.alignment = left
        else:
            cell.value = val
            cell.font = Font(size=9)
            cell.alignment = left if j != 1 else center
        cell.border = border

ws3.freeze_panes = "A3"

# ═══════════════════════════════════════════════
# SHEET 4 — DETAIL GEDUNG K
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet("GEDUNG K")
ws4.sheet_properties.tabColor = "843C0C"
ws4.column_dimensions["A"].width = 8
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 45
ws4.column_dimensions["D"].width = 20
ws4.column_dimensions["E"].width = 12

ws4.merge_cells("A1:E1")
ws4["A1"].value = "GEDUNG K — Detail File DED"
ws4["A1"].font = Font(bold=True, size=12, color="843C0C")

for col, h in enumerate(h2, 1):
    c = ws4.cell(2, col)
    c.value = h
    c.fill = PatternFill("solid", fgColor="843C0C")
    c.font = Font(color="FFFFFF", bold=True, size=10)
    c.alignment = center
    c.border = border

k_data = [
    ("1", "ARSITEKTUR", "1. ARSITEKTUR/", "Folder", "20+ DWG"),
    ("", "", "00. DAFTAR GAMBAR / SIMBOL", "Folder", ""),
    ("", "", "01. DENAH/", "Folder", "DWG"),
    ("", "", "02. TAMPAK/", "Folder", ""),
    ("", "", "03. POTONGAN/", "Folder", ""),
    ("", "", "04. POLA LANTAI/", "Folder", ""),
    ("", "", "05. POLA PLAFOND/", "Folder", ""),
    ("", "", "06. TITIK LAMPU/", "Folder", ""),
    ("", "", "07. DETAIL TANGGA/", "Folder", ""),
    ("", "", "08. RENCANA KUSEN/", "Folder", ""),
    ("", "", "09. DETAIL TOILET/", "Folder", ""),
    ("", "", "  - Ars PDF (di folder Ars)", "", ""),
    ("2", "STRUKTUR", "3. STRUKTUR/STR GED K 260330/", "Folder", "1 DWG + 17 PDF"),
    ("", "", "2.STR GEDUNG K.dwg", "DWG", "Struktur utama"),
    ("", "", "DED STR GED K 260328.pdf", "PDF", "DED konsolidated"),
    ("", "", "SD-01.pdf ~ SD-05.pdf", "PDF", "Standar Detail"),
    ("", "", "STR-00.pdf ~ STR-16.pdf", "PDF", "Detail struktur per bagian"),
    ("3", "MEP", "4. MEP/MEP GED K 260330/", "Folder", "CAD + PDF"),
    ("", "", "BIN - CAD & PDF (Elektrikal)/", "Folder", "30+ PDF"),
    ("", "", "  - #0 UMUM, #1 DIAGRAM PANEL, #2 INSTALASI, #3 DETAIL", "", ""),
    ("", "", "GEDUNG K CAD&PDF (PL&PK)/", "Folder", "Plumbing + Pemadam"),
    ("", "", "  - PK: SISTEM, INSTALASI, DETAIL", "", ""),
    ("", "", "  - PL: SISTEM, INSTALASI, DETAIL", "", ""),
    ("", "", "EC (Elektronik)/", "Folder", "CCTV, Fire Alarm, dll"),
    ("", "", "TDG (Lift)/", "Folder", "Detail lift"),
    ("", "", "VAC (AC)/", "Folder", "AC / Ventilasi"),
    ("", "", "DED MEP GED K 260328.pdf", "PDF", "DED konsolidated"),
    ("", "", "EC GED K 260328.pdf", "PDF", "Elektronik"),
    ("", "", "EL GED K 260328.pdf", "PDF", "Elektrikal"),
    ("", "", "PK GED K 260328.pdf", "PDF", "Pemadam"),
    ("", "", "PL GED K 260328.pdf", "PDF", "Plumbing"),
    ("", "", "TDG GED K 260328.pdf", "PDF", "Lift"),
    ("", "", "VAC GED K 260328.pdf", "PDF", "AC"),
    ("", "", "BIN GD.K PLPK R 21 May/", "Folder", "Revisi May 2026"),
    ("4", "INTERIOR", "2. INTERIOR/", "Folder", "PDF + XLSX"),
    ("", "", "1. BUILT IN FURNITURE/GEDUNG K_BUILT IN FURNITUR_260327.pdf", "PDF", ""),
    ("", "", "2. LOOSE FURNITURE/GEDUNG K_LOOSE FURNITUR_260327.pdf", "PDF", ""),
    ("", "", "INTERIOR GEDUNG K.xlsx", "XLSX", ""),
    ("5", "RAB", "5. RAB/", "Folder", "3 XLSX + 1 PDF"),
    ("", "", "260326_RAB PEKERJAAN FISIK PEMBANGUNAN GEDUNG K BIN.xlsx", "XLSX", ""),
    ("", "", "260326_RAB UNIT PENGADAAN PEMBANGUNAN GEDUNG K BIN.xlsx", "XLSX", ""),
    ("", "", "4. RAB GEDUNG K.pdf", "PDF", ""),
    ("", "", "INTERIOR GEDUNG K.xlsx", "XLSX", ""),
    ("6", "OUTLINE SPEK", "20260513 OUTLINE SPEK GEDUNG K.pdf", "PDF", "May 2026"),
]

for i, row_data in enumerate(k_data, 3):
    ws4.row_dimensions[i].height = 18
    for j, val in enumerate(row_data, 1):
        cell = ws4.cell(i, j)
        if j == 2 and val:
            cell.value = val
            cell.fill = PatternFill("solid", fgColor="A5520E")
            cell.font = Font(color="FFFFFF", bold=True, size=9)
            cell.alignment = left
        elif val.endswith("/"):
            cell.value = val
            cell.font = Font(bold=True, size=9, color="843C0C")
            cell.alignment = left
        else:
            cell.value = val
            cell.font = Font(size=9)
            cell.alignment = left if j != 1 else center
        cell.border = border

ws4.freeze_panes = "A3"

# ═══════════════════════════════════════════════
# SHEET 5 — INTERIOR SHARED
# ═══════════════════════════════════════════════
ws5 = wb.create_sheet("INTERIOR")
ws5.sheet_properties.tabColor = "7030A0"
ws5.column_dimensions["A"].width = 8
ws5.column_dimensions["B"].width = 18
ws5.column_dimensions["C"].width = 50
ws5.column_dimensions["D"].width = 15

ws5.merge_cells("A1:D1")
ws5["A1"].value = "INTERIOR — Detail File (Folder Terpisah)"
ws5["A1"].font = Font(bold=True, size=12, color="7030A0")

for col, h in enumerate(["NO", "GEDUNG", "NAMA FILE", "TIPE"], 1):
    c = ws5.cell(2, col)
    c.value = h
    c.fill = PatternFill("solid", fgColor="7030A0")
    c.font = Font(color="FFFFFF", bold=True, size=10)
    c.alignment = center
    c.border = border

int_data = [
    ("1", "GEDUNG B", "20260429 DED INTERIOR OFFICE B.pdf", "PDF"),
    ("2", "GEDUNG D", "01. CAD/ (18 file — Cover, Scope, Drawing List, Denah, Detail Ruangan)", "CAD"),
    ("", "", "  Ruangan: LOBBY, STAFF, DIREKTUR, DEPUTI, ESELON 3, RAPAT BESAR", "", ""),
    ("", "", "  RAPAT DIREKTORAT, STAFF PERCETAKAN, AHLI MADYA, TUNGGU, TELECONFERENCE", "", ""),
    ("3", "GEDUNG D", "02. PDF/ (sama — 18 file PDF)", "PDF"),
    ("4", "GEDUNG D", "03. RKS & OUTLINE SPEK/RKS INTERIOR BIN - GEDUNG D.pdf", "PDF"),
    ("5", "GEDUNG D", "BQ Interior gedung D.xlsx", "XLSX"),
    ("6", "GEDUNG K", "1. BUILT IN FURNITURE/GEDUNG K_BUILT IN FURNITUR_260327.pdf", "PDF"),
    ("7", "GEDUNG K", "2. LOOSE FURNITURE/GEDUNG K_LOOSE FURNITUR_260327.pdf", "PDF"),
    ("8", "GEDUNG K", "INTERIOR GEDUNG K.xlsx", "XLSX"),
]

for i, row_data in enumerate(int_data, 3):
    ws5.row_dimensions[i].height = 18
    for j, val in enumerate(row_data, 1):
        cell = ws5.cell(i, j)
        if j == 2 and val in ("GEDUNG B", "GEDUNG D", "GEDUNG K"):
            cell.value = val
            cell.fill = PatternFill("solid", fgColor="9966CC")
            cell.font = Font(color="FFFFFF", bold=True, size=9)
        else:
            cell.value = val
            cell.font = Font(size=9)
        cell.alignment = left if j != 1 else center
        cell.border = border

ws5.freeze_panes = "A3"

# Save
out = r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx"
wb.save(out)
print(f"✅ Saved: {out}")
