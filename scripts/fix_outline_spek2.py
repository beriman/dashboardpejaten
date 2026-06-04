#!/usr/bin/env python3
"""Fix 2 masalah: 1) Pindahkan OUTLINE SPEK di GEDUNG D ke posisi benar, 2) Isi kolom OUTLINE SPEK di SUMMARY."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.load_workbook(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")

# ════════════════════════════════════════
# FIX 1: Pindahkan OUTLINE SPEK di GEDUNG D
# ════════════════════════════════════════
ws_d = wb["GEDUNG D"]

# Cari baris OUTLINE SPEK dan semua sub-itemnya
outline_start = None
outline_end = None
for row_num in range(1, ws_d.max_row + 1):
    if ws_d.cell(row_num, 2).value == "OUTLINE SPEK":
        outline_start = row_num
        # Cari akhir sub-item
        for sub in range(row_num + 1, ws_d.max_row + 1):
            if ws_d.cell(sub, 2).value is None and ws_d.cell(sub, 3).value and str(ws_d.cell(sub, 3).value).startswith("  "):
                outline_end = sub
            else:
                break
        break

print(f"OUTLINE SPEK: rows {outline_start}-{outline_end}")

# Cari posisi insert yang benar (sebelum RAB)
insert_before = None
for row_num in range(1, ws_d.max_row + 1):
    if ws_d.cell(row_num, 2).value == "RAB":
        insert_before = row_num
        break

print(f"Insert before row {insert_before} (RAB)")

if outline_start and outline_end and insert_before:
    # Copy data OUTLINE SPEK
    outline_data = []
    for row_num in range(outline_start, outline_end + 1):
        row_data = []
        for col in range(1, 6):
            cell = ws_d.cell(row_num, col)
            row_data.append((cell.value, cell.fill, cell.font, cell.alignment, cell.border))
        outline_data.append(row_data)
    
    # Hapus baris lama (dari bawah ke atas)
    for row_num in range(outline_end, outline_start - 1, -1):
        ws_d.delete_rows(row_num)
    
    # Adjust insert position after deletion
    if insert_before > outline_end:
        insert_before -= (outline_end - outline_start + 1)
    
    # Insert di posisi baru
    for i, row_data in enumerate(outline_data):
        row_num = insert_before + i
        ws_d.insert_rows(row_num)
        ws_d.row_dimensions[row_num].height = 18
        for col, (val, fill, font, align, border) in enumerate(row_data, 1):
            cell = ws_d.cell(row_num, col)
            cell.value = val
            if fill and fill.fill_type != 'none':
                cell.fill = fill
            if font:
                cell.font = font
            if align:
                cell.alignment = align
            if border:
                cell.border = border

# Update nomor
for row_num in range(1, ws_d.max_row + 1):
    cell_b = ws_d.cell(row_num, 2).value
    if cell_b == "OUTLINE SPEK":
        ws_d.cell(row_num, 1).value = "7"
    elif cell_b == "RAB":
        ws_d.cell(row_num, 1).value = "8"
    elif cell_b == "LAINNYA":
        ws_d.cell(row_num, 1).value = "9"

# ════════════════════════════════════════
# FIX 2: Isi kolom OUTLINE SPEK di SUMMARY
# ════════════════════════════════════════
ws_sum = wb["SUMMARY"]

thin = Side(style="thin", color="BFBFBF")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
green_fill = PatternFill("solid", fgColor="C6EFCE")
green_font = Font(color="006100", size=9, bold=True)
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
yellow_font = Font(color="9C5700", size=9)
na_fill = PatternFill("solid", fgColor="F2F2F2")
na_font = Font(color="7F7F7F", size=9, italic=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data outline spek per disiplin (berdasarkan file yang ditemukan)
outline_map = {
    "ARSITEKTUR": "ADA",       # Gedung B: RKS GEDUNG B (DEPUTI-I).pdf, Gedung D: ada, Gedung K: ada
    "STRUKTUR": "BELUM",       # Gedung B: tidak ada outline terpisah, Gedung D: ada di Drive, Gedung K: tidak ada
    "MEP (MEKANIKAL ELEKTRONIK)": "BELUM",  # Gedung B: tidak ada, Gedung D: ada di Drive, Gedung K: tidak ada
    "INTERIOR": "BELUM",       # Gedung B: tidak ada, Gedung D: ada, Gedung K: tidak ada
    "SITE DEVELOPMENT": "BELUM",
    "OUTLINE SPESIFIKASI": "ADA",  # Semua gedung ada
}

for row_num in range(4, ws_sum.max_row + 1):
    disc = ws_sum.cell(row_num, 1).value
    cell = ws_sum.cell(row_num, 5)
    
    if disc and not str(disc).startswith("  "):
        # Header disiplin
        disc_key = str(disc).strip()
        status = outline_map.get(disc_key)
        if status == "ADA":
            cell.value = "ADA"; cell.fill = green_fill; cell.font = green_font
        elif status == "BELUM":
            cell.value = "BELUM"; cell.fill = yellow_fill; cell.font = yellow_font
        else:
            cell.value = "—"; cell.fill = na_fill; cell.font = na_font
    else:
        # Sub-item
        cell.value = "—"; cell.fill = na_fill; cell.font = na_font
    cell.alignment = center; cell.border = bdr

wb.save(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")
print("✅ Fixed: OUTLINE SPEK dipindahkan + SUMMARY kolom 5 diisi")
