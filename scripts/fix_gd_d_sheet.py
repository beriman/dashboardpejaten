#!/usr/bin/env python3
"""Rebuild sheet GEDUNG D dengan format yang sejajar dengan GEDUNG B dan K."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.load_workbook(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")

# ── Styles (sama dengan sheet B dan K) ──
green_fill = PatternFill("solid", fgColor="C6EFCE")
green_font = Font(color="006100", size=9)
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
yellow_font = Font(color="9C5700", size=9)
red_fill = PatternFill("solid", fgColor="FFC7CE")
red_font = Font(color="9C0006", size=9)
na_fill = PatternFill("solid", fgColor="D9D9D9")
na_font = Font(color="7F7F7F", size=9, italic=True)
sub_fill_gd = PatternFill("solid", fgColor="548235")
sub_font = Font(color="FFFFFF", bold=True, size=9)
hdr_fill_gd = PatternFill("solid", fgColor="375623")
hdr_font = Font(color="FFFFFF", bold=True, size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, val):
    cell.value = val; cell.fill = hdr_fill_gd; cell.font = hdr_font
    cell.alignment = center; cell.border = border

def cat(cell, val):
    cell.value = val; cell.fill = sub_fill_gd; cell.font = sub_font
    cell.alignment = left; cell.border = border

def dat(cell, val, fill=None, font=None, align=left):
    cell.value = val
    if fill: cell.fill = fill
    if font: cell.font = font
    cell.alignment = align; cell.border = border

def dat_yn(cell, val):
    cell.value = val; cell.alignment = center; cell.border = border
    if "ADA" in val and "TIDAK" not in val:
        cell.fill = green_fill; cell.font = green_font
    elif "BELUM" in val or "TIDAK" in val:
        cell.fill = red_fill; cell.font = red_font
    elif "SUBMITTED" in val or "Proses" in val:
        cell.fill = yellow_fill; cell.font = yellow_font
    else:
        cell.fill = na_fill; cell.font = na_font

# Hapus sheet GEDUNG D lama
if "GEDUNG D" in wb.sheetnames:
    del wb["GEDUNG D"]

ws = wb.create_sheet("GEDUNG D")
ws.sheet_properties.tabColor = "375623"
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 45
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 12

ws.merge_cells("A1:E1")
ws["A1"].value = "GEDUNG D — Detail File DED"
ws["A1"].font = Font(bold=True, size=12, color="375623")

for col, h in enumerate(["NO", "KATEGORI", "NAMA FILE / FOLDER", "TIPE", "KET"], 1):
    hdr(ws.cell(2, col), h)

# ═══ DATA — format sejajar: setiap baris punya NO + KATEGORI ═══
d_data = [
    # --- ARSITEKTUR ---
    ("1", "ARSITEKTUR", "01. ARSITEKTUR-20260504T035629Z-3-001/", "Folder", "144 file, lengkap"),
    ("", "", "00. EKSISTING/Gedung D EKSISTING - 10-01-2018.dwg", "DWG", "Data existing"),
    ("", "", "01. CAD/00. UMUM/", "Folder", "COVER, SYMBOL, PERSPEKTIF, Masterplan"),
    ("", "", "01. CAD/01. SITE PLAN & BLOCK PLAN/", "Folder", "AK0101-AK0102"),
    ("", "", "01. CAD/02. DENAH GENERAL/", "Folder", "LT1-4 + ATAP (6 DWG)"),
    ("", "", "01. CAD/02. DENAH PARSIAL/", "Folder", "Parsial LT1-ATAP (12 DWG)"),
    ("", "", "01. CAD/03. TAMPAK & POTONGAN/", "Folder", "Tampak 1,2 + Potong 1,2"),
    ("", "", "01. CAD/04. PARSIAL FASAD/", "Folder", "Selatan, Timur, Utara, Barat"),
    ("", "", "01. CAD/05. DETAIL FASAD/", "Folder", "Perforated, Gawangan, Jendela"),
    ("", "", "01. CAD/06. DETAIL CORE/", "Folder", "Denah, Pola Lantai, Pola Plafon"),
    ("", "", "01. CAD/07. DETAIL TANGGA/", "Folder", "AK0701-AK0702 (8 DWG)"),
    ("", "", "01. CAD/08. DETAIL TOILET/", "Folder", "Denah, Potongan, Legend"),
    ("", "", "01. CAD/09. POLA LANTAI/", "Folder", "LT1-ATAP parsial (15 DWG+PDF)"),
    ("", "", "01. CAD/10. POLA PLAFON/", "Folder", "LT1-ATAP parsial (15 DWG)"),
    ("", "", "01. CAD/11. SKEMA KUSEN/", "Folder", "AK1101-AK1108 (8 DWG)"),
    ("", "", "01. CAD/12. DETAIL KHUSUS/", "Folder", ""),
    ("", "", "01. CAD/13. DETAIL MEP/", "Folder", ""),
    ("", "", "01. CAD/14. SARANA LUAR/", "Folder", "Drop Off, Planter"),
    ("", "", "01. CAD/15. DETAIL STANDAR/", "Folder", ""),
    ("", "", "02. PDF/", "Folder", "PDF consolidated"),
    ("", "", "03. RKS & OUTLINE SPEK/", "Folder", "RKS Arsitektur"),
    ("", "", "05. ARSITEKTUR_Drive/", "Folder", "52 file dari Drive — revisi terbaru"),
    # --- INTERIOR ---
    ("2", "INTERIOR", "02. INTERIOR/", "Folder", "41 file lengkap"),
    ("", "", "01. CAD/INT 0 - INFORMASI UMUM/", "Folder", "Cover, Scope, Drawing List"),
    ("", "", "01. CAD/INT 1 - DENAH GENERAL/", "Folder", "Denah interior"),
    ("", "", "01. CAD/INT 2 - DETAIL RUANGAN/", "Folder", "14 ruangan detail"),
    ("", "", "02. PDF/", "Folder", "18 file PDF"),
    ("", "", "03. RKS & OUTLINE SPEK/RKS INTERIOR BIN - GEDUNG D.pdf", "PDF", ""),
    ("", "", "BQ Interior gedung D.xlsx", "XLSX", ""),
    ("", "", "06. INTERIOR_Drive/", "Folder", "109 file dari Drive — lebih lengkap"),
    # --- STRUKTUR ---
    ("3", "STRUKTUR", "03. STRUKTUR-20260504T035630Z-3-001/", "Folder", "CAD + PDF + RKS"),
    ("", "", "03. STRUKTUR_Drive/", "Folder", "19 file dari Drive"),
    ("", "", "  X_KOP_Gedung D_Rev1.dwg", "DWG", "Kop terbaru"),
    # --- INFRASTRUKTUR ---
    ("4", "INFRASTRUKTUR", "04. INFRASTRUKTUR/", "Folder", "CAD + PDF + RKS"),
    ("", "", "04. INFRASTRUKTUR-20260504T035630Z-3-001/", "Folder", "Mirror dari Drive"),
    ("", "", "08. INFRASTRUKTUR_Drive/", "Folder", "1 file dari Drive"),
    # --- MEKANIKAL ---
    ("5", "MEKANIKAL", "05. MEKANIKAL-20260504T035631Z-3-001/", "Folder", "74 file lengkap"),
    ("", "", "01. CAD/01. AC/", "Folder", "DWG"),
    ("", "", "01. CAD/02. LIFT/", "Folder", "DWG"),
    ("", "", "01. CAD/03. PLUMBING/", "Folder", "DWG"),
    ("", "", "01. CAD/04. HYDRANT/", "Folder", "DWG"),
    ("", "", "01. CAD/05. SPRINKLER/", "Folder", "DWG"),
    ("", "", "02. PDF/DED MEKANIKAL GEDUNG KANTOR D.pdf", "PDF", "Konsolidated"),
    ("", "", "03. RKS & OUTLINE SPEK/RKS & OUTLINE SPEK MEKANIKAL.pdf", "PDF", ""),
    # --- ELEKTRIKAL ---
    ("6", "ELEKTRIKAL", "06. ELEKTRIKAL/", "Folder", "86 file lengkap"),
    ("", "", "06. ELEKTRIKAL-20260504T035637Z-3-001/", "Folder", "Mirror dari Drive"),
    ("", "", "10. ELEKTRIKAL_Drive/", "Folder", "86 file dari Drive"),
    ("", "", "01. CAD/LAK/ (9 subfolder)", "Folder", "Daya: Siteplan, System, Pencahayaan, Kotak Kontak, Panel, Genset, Petir, Kabel Tray, Detail"),
    ("", "", "01. CAD/LAL/ (5 subfolder)", "Folder", "Low Voltage: Fire Alarm, Data/Telpon, Security, Conference, Daftar Gambar"),
    ("", "", "02. PDF/DED ELEKTRIKAL GEDUNG KANTOR D.pdf", "PDF", "Konsolidated"),
    ("", "", "03. RKS & OUTLINE SPEK/RKS & OUTLINE SPEK ELEKTRIKAL.pdf", "PDF", ""),
    # --- OUTLINE SPEK ---
    ("7", "OUTLINE SPEK", "11. OUTLINE_SPEK_Drive/", "Folder", "11 file dari Drive"),
    # --- RKS ---
    ("8", "RKS", "12. RKS_Drive/", "Folder", "1 file dari Drive"),
    # --- RAB ---
    ("9", "RAB", "2026.05.22 RAB GD. KANTOR BIN (Konstruksi) - konsinyering JAKON.xlsx", "XLSX", ""),
    ("", "", "2026.05.22 RAB GD. KANTOR BIN (Pengadaan) - Konsinyering JAKON.xlsx", "XLSX", ""),
    # --- LAINNYA ---
    ("10", "LAINNYA", "2026.05.11_Gedung D_Revisi GWT.dwg", "DWG", "Revisi GWT"),
    ("", "", "STP Biofilter.dwg", "DWG", ""),
]

# ═══ Write data ═══
current_cat = None
for i, row_data in enumerate(d_data, 3):
    ws.row_dimensions[i].height = 18
    no, category, name, ftype, ket = row_data

    # Column A: NO
    cell_a = ws.cell(i, 1)
    cell_a.value = no
    cell_a.alignment = center
    cell_a.border = border
    cell_a.font = Font(size=9)

    # Column B: KATEGORI — hanya isi jika kategori baru (sejajar dengan B dan K)
    cell_b = ws.cell(i, 2)
    if category and category != current_cat:
        cat(cell_b, category)
        current_cat = category
    else:
        cell_b.value = None
        cell_b.border = border

    # Column C: NAMA FILE
    cell_c = ws.cell(i, 3)
    if name.endswith("/"):
        cell_c.value = name
        cell_c.font = Font(bold=True, size=9, color="2E75B6")
    elif name.startswith("  "):
        cell_c.value = name
        cell_c.font = Font(size=8, color="444444")
    else:
        cell_c.value = name
        cell_c.font = Font(size=9)
    cell_c.alignment = left
    cell_c.border = border

    # Column D: TIPE
    cell_d = ws.cell(i, 4)
    cell_d.value = ftype
    cell_d.alignment = center
    cell_d.border = border
    cell_d.font = Font(size=9)

    # Column E: KET
    cell_e = ws.cell(i, 5)
    cell_e.value = ket
    cell_e.alignment = left
    cell_e.border = border
    cell_e.font = Font(size=9)

ws.freeze_panes = "A3"

# ═══ Update SUMMARY sheet kolom Gedung D ═══
ws_sum = wb["SUMMARY"]
# Mapping baris SUMMARY ke status Gedung D
# Berdasarkan struktur SUMMARY yang ada
sum_updates = {
    6:  "Gedung D",           # Header baris
    7:  "ADA",   # Denah/Site Plan
    8:  "ADA",   # Tampak/Potongan
    9:  "ADA",   # Pola Lantai
    10: "ADA",   # Plafond
    11: "ADA",   # Kusen/Pintu
    12: "ADA",   # Toilet Detail
    13: "ADA",   # Tangga Detail
    14: "ADA",   # Fasad Detail
    15: "ADA",   # Schedule Material
    16: "ADA",   # Ars PDF
    17: "~65 DWG",  # DWG count
    19: "ADA",   # Pondasi
    20: "ADA",   # Sloof/Balok
    21: "ADA",   # Kolom
    22: "ADA",   # Plat Lantai
    23: "ADA",   # Portal/Rangka Atap
    24: "ADA",   # Standar Detail
    25: "ADA",   # Struktur PDF
    26: "~20 DWG",  # Struktur DWG
    28: "ADA",   # Plumbing
    29: "ADA",   # Pemadam
    30: "ADA",   # AC/VAC
    31: "ADA",   # Elektrikal
    32: "ADA",   # Elektronik
    33: "ADA",   # Lift
    34: "ADA",   # MEP PDF
    35: "~80+ DWG",  # MEP DWG
    37: "ADA",   # Denah Interior
    38: "ADA",   # Detail Ruangan
    39: "ADA",   # Built-in Furniture
    40: "ADA",   # Loose Furniture
    41: "ADA",   # BOQ/RKS Interior
    44: "ADA",   # Site Plan/Grading
    45: "ADA",   # Site Dev PDF
    47: "ADA",   # RKS Arsitektur
    48: "ADA",   # RKS Interior
    49: "BELUM", # RKS Mekanikal
    50: "ADA",   # RKS Elektrikal
    52: "ADA",   # RAB Pekerjaan Fisik
    53: "ADA",   # RAB Interior
    54: "ADA",   # BOQ MEP
    56: "TIDAK ADA",  # Laporan Pendahuluan
    57: "TIDAK ADA",  # Laporan Antara
    58: "TIDAK ADA",  # Laporan Akhir Draft
    59: "TIDAK ADA",  # Laporan Akhir Final
    61: "ADA",   # Outline Spek
}

for row_num, val in sum_updates.items():
    cell = ws_sum.cell(row_num, 3)  # Column C = Gedung D
    cell.value = val
    if val == "ADA":
        cell.fill = green_fill; cell.font = green_font
    elif val == "TIDAK ADA":
        cell.fill = red_fill; cell.font = red_font
    elif val == "BELUM":
        cell.fill = yellow_fill; cell.font = yellow_font
    else:
        cell.fill = na_fill; cell.font = na_font
    cell.alignment = center
    cell.border = border

wb.save(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")
print("✅ GEDUNG D sheet rebuilt — format sejajar dengan B dan K")
