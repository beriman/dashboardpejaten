#!/usr/bin/env python3
"""Rebuild SUMMARY sheet dengan data akurat + format rapi."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.load_workbook(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")

# ═══ Styles ═══
hdr_fill  = PatternFill("solid", fgColor="1F4E79")
hdr_font  = Font(color="FFFFFF", bold=True, size=10)
disc_fill = PatternFill("solid", fgColor="2E75B6")
disc_font = Font(color="FFFFFF", bold=True, size=9)
sub_fill  = PatternFill("solid", fgColor="D6DCE4")
sub_font  = Font(bold=True, size=9, color="1F4E79")
green_fill = PatternFill("solid", fgColor="C6EFCE")
green_font = Font(color="006100", size=9, bold=True)
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
yellow_font = Font(color="9C5700", size=9)
red_fill  = PatternFill("solid", fgColor="FFC7CE")
red_font  = Font(color="9C0006", size=9)
na_fill   = PatternFill("solid", fgColor="F2F2F2")
na_font   = Font(color="7F7F7F", size=9, italic=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(cell, val, fill, font, align=center):
    cell.value = val; cell.fill = fill; cell.font = font
    cell.alignment = align; cell.border = bdr

# ═══ Helper: status → (fill, font) ═══
def sf(status):
    s = status.strip().upper()
    if s in ("ADA", "✅ ADA"):
        return green_fill, green_font
    elif "SUBMITTED" in s or "PROSES" in s:
        return yellow_fill, yellow_font
    elif s in ("BELUM", "PARTIAL", "SEBAGIAN"):
        return yellow_fill, yellow_font
    elif s in ("TIDAK ADA", "TIDAK", "N/A", ""):
        return na_fill, na_font
    else:
        return na_fill, na_font

# ════════════════════════════════════════
# SHEET 1 — SUMMARY
# ════════════════════════════════════════
ws = wb["SUMMARY"]

# Clear old data (keep row 1-3 header)
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        cell.value = None
        cell.fill = PatternFill()
        cell.font = Font(size=9)
        cell.border = bdr

ws.column_dimensions["A"].width = 28
for col in "BCDEFGHI":
    ws.column_dimensions[col].width = 16

# ═══ Title ═══
ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 16
ws.row_dimensions[3].height = 40

ws.merge_cells("A1:I1")
cell_style(ws["A1"], "DED CHECKLIST — PROJECT PEJATEN (BIN RENOVASI)",
           PatternFill(), Font(bold=True, size=14, color="1F4E79"), center)

ws.merge_cells("A2:I2")
cell_style(ws["A2"],
           "Sumber: Google Drive 'FILE DIRECTORY BIN 2' + scan folder lokal — Update: 04 Juni 2026",
           PatternFill(), Font(size=9, color="808080", italic=True), center)

# ═══ Column headers row 3 ═══
headers = ["DISIPLIN / DOKUMEN", "GEDUNG B", "GEDUNG D", "GEDUNG K",
            "INTERIOR (B/D/K)", "RKS", "RAB", "LAPORAN", "OUTLINE SPEK"]
for c, h in enumerate(headers, 1):
    cell_style(ws.cell(3, c), h, hdr_fill, hdr_font)

# ═══ Data rows ═══
# Format: (DISIPLIN, GEDUNG B, GEDUNG D, GEDUNG K, INTERIOR, RKS, RAB, LAPORAN, OUTLINE_SPEK)

rows = [
    # ── ARSITEKTUR ──
    ("ARSITEKTUR", "", "", "", "", "", "", "", ""),
    ("  Denah / Site Plan",         "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Tampak & Potongan",         "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Pola Lantai",               "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Plafond",                   "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Kusen / Pintu",             "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Toilet Detail",             "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Tangga Detail",             "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Fasad Detail",              "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Schedule Material",         "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Ars PDF konsolidated",      "ADA", "BELUM","BELUM","",  "",    "",    "",  ""),
    ("  Ars DWG count",        "~15 DWG","~65 DWG","~20 DWG","", "",    "",    "",  ""),
    # ── STRUKTUR ──
    ("STRUKTUR", "", "", "", "", "", "", "", ""),
    ("  Pondasi",                   "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Sloof / Balok",             "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Kolom",                     "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Plat Lantai",               "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Portal / Rangka Atap",      "ADA", "BELUM","ADA", "",  "",    "",    "",  ""),
    ("  Standar Detail (SD-01~06)",  "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Struktur PDF",              "ADA", "BELUM","ADA","",   "",    "",    "",  ""),
    ("  Struktur DWG count",   "~5 DWG","~20 DWG","~3 DWG","",  "",    "",    "",  ""),
    # ── MEP ──
    ("MEP (Mekanikal Elektronik)", "", "", "", "", "", "", "", ""),
    ("  Plumbing (PL)",             "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Pemadam Kebakaran (PK)",     "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  AC / VAC",                   "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Elektrikal (LAK+LAL)",       "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
    ("  Elektronik / Low Voltage",  "BELUM","ADA", "ADA", "",  "",    "",    "",  ""),
    ("  Lift / TDG",                "BELUM","BELUM","ADA","",  "",    "",    "",  ""),
    ("  MEP PDF konsolidated",      "ADA", "BELUM","ADA","",   "",    "",    "",  ""),
    ("  MEP DWG count",       "~15 DWG","~80+ DWG","~40+ DWG","", "",   "",    "",  ""),
    # ── INTERIOR ──
    ("INTERIOR", "", "", "", "", "", "", "", ""),
    ("  Denah Interior",            "",    "",    "",    "ADA (D,K)", "", "",   "",  ""),
    ("  Detail Ruangan",            "",    "",    "",    "ADA (D)",  "", "",   "",  ""),
    ("  Built-in Furniture",        "",    "",    "",    "ADA (K)",  "", "",   "",  ""),
    ("  Loose Furniture",           "",    "",    "",    "ADA (K)",  "", "",   "",  ""),
    ("  RKS Interior",             "",    "",    "",    "ADA (D)",  "", "",   "",  ""),
    ("  BOQ Interior",             "",    "",    "",    "ADA (D,K)","", "",   "",  ""),
    # ── SITE DEVELOPMENT ──
    ("SITE DEVELOPMENT", "", "", "", "", "", "", "", ""),
    ("  Site Plan / Grading",        "ADA", "ADA", "BELUM","",  "",    "",    "",  ""),
    ("  PDF",                        "ADA", "ADA", "BELUM","",  "",    "",    "",  ""),
    # ── RKS ──
    ("RKS", "", "", "", "", "", "", "", ""),
    ("  RKS Arsitektur",            "BELUM","BELUM","BELUM","", "ADA (B)", "",  "",  ""),
    ("  RKS Interior",              "",    "",    "",    "",   "ADA (D)", "",  "",  ""),
    ("  RKS Mekanikal",             "BELUM","BELUM","BELUM","", "",    "",    "",  ""),
    ("  RKS Elektrikal",            "",    "ADA",  "BELUM","", "",    "",    "",  ""),
    ("  RKS Infrastruktur",         "",    "ADA",  "BELUM","", "",    "",    "",  ""),
    # ── RAB ──
    ("RAB / BOQ", "", "", "", "", "", "", "", ""),
    ("  RAB Pekerjaan Fisik",        "ADA", "",    "ADA", "",   "",    "",    "",  ""),
    ("  RAB Pengadaan / Konstruksi", "",   "ADA",  "",    "",   "",    "",    "",  ""),
    ("  RAB Interior",              "",    "",     "ADA", "",   "",    "",    "",  ""),
    ("  BOQ Interior",              "",    "",     "",    "ADA (D,K)","", "",  "",  ""),
    # ── LAPORAN ──
    ("LAPORAN", "", "", "", "", "", "", "", ""),
    ("  Laporan Pendahuluan",        "ADA", "BELUM","BELUM","", "",    "",    "",  ""),
    ("  Laporan Antara",             "ADA", "BELUM","BELUM","", "",    "",    "",  ""),
    ("  Draft Laporan Akhir",        "ADA", "BELUM","BELUM","", "",    "",    "",  ""),
    ("  Laporan Akhir (Final)",      "ADA", "BELUM","BELUM","", "",    "",    "",  ""),
    # ── OUTLINE SPESIFIKASI ──
    ("OUTLINE SPESIFIKASI", "", "", "", "", "", "", "", ""),
    ("  Outline Spek Gedung",        "ADA", "ADA", "ADA", "",   "",    "",    "",  ""),
]

current_disc = None
for i, row_data in enumerate(rows, 4):
    ws.row_dimensions[i].height = 20
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(i, c)
        cell.border = bdr

        if c == 1:  # Discipline column
            # Detect discipline header (no indent)
            if not val.startswith("  "):
                current_disc = val
                cell.value = val
                cell.fill = disc_fill
                cell.font = disc_font
                cell.alignment = left
            else:
                cell.value = val
                cell.font = Font(size=9)
                cell.alignment = left
        else:  # Status columns
            if not val or val == "—":
                cell.value = "—"
                cell.fill = na_fill
                cell.font = na_font
                cell.alignment = center
            else:
                fill, font = sf(val)
                cell.value = val
                cell.fill = fill
                cell.font = font
                cell.alignment = center

ws.freeze_panes = "B4"

wb.save(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")
print("✅ SUMMARY rebuilt — format rapi, data akurat")
print(f"   Total baris data: {len(rows)}")
