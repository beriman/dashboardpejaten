#!/usr/bin/env python3
"""Fix: pindahkan OUTLINE SPEK ke posisi benar di GEDUNG D + isi kolom OUTLINE SPEK di SUMMARY."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy

wb = openpyxl.load_workbook(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")

# ═══ Fix 1: Hapus baris OUTLINE SPEK yang salah posisi di GEDUNG D ═══
ws_d = wb["GEDUNG D"]

# Cari baris "7" (OUTLINE SPEK) yang salah posisi
rows_to_move = []
for row_num in range(1, ws_d.max_row + 1):
    cell_b = ws_d.cell(row_num, 2).value
    if cell_b == "OUTLINE SPEK":
        # Simpan data baris ini dan semua baris sub-item setelahnya
        rows_to_move.append(row_num)
        # Cari baris sub-item (baris kosong di kolom B setelah ini)
        for sub_row in range(row_num + 1, ws_d.max_row + 1):
            if ws_d.cell(sub_row, 2).value is None and ws_d.cell(sub_row, 3).value and str(ws_d.cell(sub_row, 3).value).startswith("  "):
                rows_to_move.append(sub_row)
            else:
                break
        break

print(f"Rows to move: {rows_to_move}")

# Hapus baris yang salah posisi (dari bawah ke atas)
for row_num in reversed(rows_to_move):
    ws_d.delete_rows(row_num)

# Sekarang cari posisi yang benar untuk OUTLINE SPEK (setelah MEP/ELEKTRIKAL, sebelum RAB)
insert_after = None
for row_num in range(1, ws_d.max_row + 1):
    cell_b = ws_d.cell(row_num, 2).value
    if cell_b in ("ELEKTRIKAL", "MEKANIKAL"):
        insert_after = row_num
    elif cell_b == "RAB":
        break

# Cari lagi posisi yang benar
insert_pos = None
for row_num in range(ws_d.max_row, 0, -1):
    cell_b = ws_d.cell(row_num, 2).value
    if cell_b == "RAB":
        insert_pos = row_num
        break

print(f"Insert OUTLINE SPEK before row {insert_pos} (RAB)")

# Data OUTLINE SPEK
outline_rows = [
    ("7", "OUTLINE SPEK", "11. OUTLINE_SPEK_Drive/", "Folder", "11 file dari Drive"),
    (None, None, "  ARS/2026.04.28 - OUTLINE SPEK ARSITEKTUR BIN - GEDUNG D.pdf", "PDF", "Arsitektur"),
    (None, None, "  ARS/2026.04.30 OUTLINE SPEK ARSITEKTUR BIN - GEDUNG D.pdf", "PDF", "Arsitektur terbaru"),
    (None, None, "  ARS/Archive/2026.04.24 - OUTLINE SPEK ARSITEKTUR GEDUNG D.pdf", "PDF", "Arsitektur lama"),
    (None, None, "  ARS/Archive/2026.04.24 - OUTLINE SPEK ARSITEKTUR GEDUNG D.xlsx", "XLSX", ""),
    (None, None, "  INT/OUTLINE SPEK INTERIOR BIN - GEDUNG D.pdf", "PDF", "Interior"),
    (None, None, "  INT/OUTLINE SPEK INTERIOR BIN_ kantor.pdf", "PDF", "Interior kantor"),
    (None, None, "  INF/RKS_Infrastruktur Gedung D.pdf", "PDF", "Infrastruktur"),
    (None, None, "  LAN/2026.03.08 OUTLINE SPEK LANSKAP - BIN.xlsx", "XLSX", "Lanskap"),
    (None, None, "  MEP/2026.04.21_OUTLINE SPEK ELEKTRIKAL GEDUNG KANTOR.pdf", "PDF", "Elektrikal"),
    (None, None, "  MEP/2026.04.21_OUTLINE SPEK MEKANIKAL GEDUNG KANTOR.pdf", "PDF", "Mekanikal"),
]

# Insert baris di posisi yang benar
thin = Side(style="thin", color="BFBFBF")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, (no, cat, name, ftype, ket) in enumerate(outline_rows):
    row_num = insert_pos + i
    ws_d.insert_rows(row_num)
    ws_d.row_dimensions[row_num].height = 18
    
    # NO
    cell_a = ws_d.cell(row_num, 1)
    cell_a.value = no; cell_a.alignment = Alignment(horizontal="center", vertical="center")
    cell_a.border = bdr; cell_a.font = Font(size=9)
    
    # KATEGORI
    cell_b = ws_d.cell(row_num, 2)
    if cat:
        cell_b.value = cat
        cell_b.fill = PatternFill("solid", fgColor="548235")
        cell_b.font = Font(color="FFFFFF", bold=True, size=9)
    cell_b.alignment = Alignment(horizontal="left", vertical="center")
    cell_b.border = bdr
    
    # NAMA FILE
    cell_c = ws_d.cell(row_num, 3)
    if name.endswith("/"):
        cell_c.value = name; cell_c.font = Font(bold=True, size=9, color="2E75B6")
    elif name.startswith("  "):
        cell_c.value = name; cell_c.font = Font(size=8, color="444444")
    else:
        cell_c.value = name; cell_c.font = Font(size=9)
    cell_c.alignment = Alignment(horizontal="left", vertical="center")
    cell_c.border = bdr
    
    # TIPE
    cell_d = ws_d.cell(row_num, 4)
    cell_d.value = ftype; cell_d.alignment = Alignment(horizontal="center", vertical="center")
    cell_d.border = bdr; cell_d.font = Font(size=9)
    
    # KET
    cell_e = ws_d.cell(row_num, 5)
    cell_e.value = ket; cell_e.alignment = Alignment(horizontal="left", vertical="center")
    cell_e.border = bdr; cell_e.font = Font(size=9)

# Update nomor RAB dan LAINNYA
for row_num in range(1, ws_d.max_row + 1):
    cell_b = ws_d.cell(row_num, 2).value
    if cell_b == "RAB":
        ws_d.cell(row_num, 1).value = "8"
    elif cell_b == "LAINNYA":
        ws_d.cell(row_num, 1).value = "9"

# ═══ Fix 2: Isi kolom OUTLINE SPEK di SUMMARY ═══
ws_sum = wb["SUMMARY"]

# Data outline spek per baris disiplin
# Cari baris disiplin dan isi kolom E
outline_spek_map = {
    "ARSITEKTUR": "ADA",
    "STRUKTUR": "BELUM",
    "MEP (MEKANIKAL ELEKTRONIK)": "BELUM",
    "INTERIOR": "BELUM",
    "SITE DEVELOPMENT": "BELUM",
    "OUTLINE SPESIFIKASI": "ADA",
}

green_fill = PatternFill("solid", fgColor="C6EFCE")
green_font = Font(color="006100", size=9, bold=True)
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
yellow_font = Font(color="9C5700", size=9)
na_fill = PatternFill("solid", fgColor="F2F2F2")
na_font = Font(color="7F7F7F", size=9, italic=True)

for row_num in range(4, ws_sum.max_row + 1):
    disc = ws_sum.cell(row_num, 1).value
    if disc and not str(disc).startswith("  "):
        # Ini baris header disiplin
        disc_key = str(disc).strip()
        status = outline_spek_map.get(disc_key, "—")
        cell = ws_sum.cell(row_num, 5)
        if status == "ADA":
            cell.value = "ADA"; cell.fill = green_fill; cell.font = green_font
        elif status == "BELUM":
            cell.value = "BELUM"; cell.fill = yellow_fill; cell.font = yellow_font
        else:
            cell.value = "—"; cell.fill = na_fill; cell.font = na_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr

wb.save(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")
print("✅ Fixed: OUTLINE SPEK dipindahkan ke posisi benar + SUMMARY updated")
