#!/usr/bin/env python3
"""Update kolom OUTLINE SPEK di SUMMARY dan update GEDUNG D sheet."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.load_workbook(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")

# ═══ Styles ═══
green_fill = PatternFill("solid", fgColor="C6EFCE")
green_font = Font(color="006100", size=9, bold=True)
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
yellow_font = Font(color="9C5700", size=9)
red_fill = PatternFill("solid", fgColor="FFC7CE")
red_font = Font(color="9C0006", size=9)
na_fill = PatternFill("solid", fgColor="F2F2F2")
na_font = Font(color="7F7F7F", size=9, italic=True)
thin = Side(style="thin", color="BFBFBF")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

def set_cell(cell, val, fill, font):
    cell.value = val; cell.fill = fill; cell.font = font
    cell.alignment = center; cell.border = bdr

# ═══ Update SUMMARY — tambah kolom OUTLINE SPEK ═══
ws = wb["SUMMARY"]

# Cek apakah sudah ada kolom ke-5 (OUTLINE SPEK)
# Dari data, kolom D = GEDUNG K. Perlu tambah kolom E = OUTLINE SPEK
# Tapi lihat struktur: user sudah punya 4 kolom (A=DISIPLIN, B=GEDUNG B, C=GEDUNG D, D=GEDUNG K)
# Perlu tambah kolom E untuk OUTLINE SPEK

# Insert kolom E
ws.insert_cols(5)
ws.column_dimensions["E"].width = 16

# Header
ws.cell(3, 5).value = "OUTLINE SPEK"
ws.cell(3, 5).fill = PatternFill("solid", fgColor="1F4E79")
ws.cell(3, 5).font = Font(color="FFFFFF", bold=True, size=10)
ws.cell(3, 5).alignment = center
ws.cell(3, 5).border = bdr

# Data outline spek per disiplin
# Baris 4 = ARSITEKTUR header
outline_data = {
    # row_num: (status, gedung_b, gedung_d, gedung_k)
    4:  ("ARSITEKTUR", "ADA", "ADA", "ADA"),      # ARSITEKTUR
    16: ("STRUKTUR", "BELUM", "ADA", "BELUM"),     # STRUKTUR
    26: ("MEP", "BELUM", "ADA", "BELUM"),          # MEP
    37: ("INTERIOR", "BELUM", "ADA", "BELUM"),     # INTERIOR
    44: ("SITE DEV", "BELUM", "ADA", "BELUM"),     # SITE DEVELOPMENT
}

# Fill semua baris dengan default "—"
for row_num in range(4, ws.max_row + 1):
    cell = ws.cell(row_num, 5)
    disc_val = ws.cell(row_num, 1).value
    if disc_val and not str(disc_val).startswith("  "):
        # Ini baris header disiplin
        cell.value = "—"; cell.fill = na_fill; cell.font = na_font
    else:
        cell.value = "—"; cell.fill = na_fill; cell.font = na_font
    cell.alignment = center; cell.border = bdr

# Override dengan data aktual
for row_num, (disc, gb, gd, gk) in outline_data.items():
    statuses = [gb, gd, gk]
    for i, status in enumerate(statuses):
        col = 2 + i  # B=2, C=3, D=4
        cell = ws.cell(row_num, col)
        s = status.upper().strip()
        if s == "ADA":
            set_cell(cell, "ADA", green_fill, green_font)
        elif s == "BELUM":
            set_cell(cell, "BELUM", yellow_fill, yellow_font)
        else:
            set_cell(cell, "—", na_fill, na_font)

# ═══ Update GEDUNG D sheet — tambah data Outline Spek ═══
ws_d = wb["GEDUNG D"]

# Cari baris terakhir
last_row = ws_d.max_row

# Tambah baris outline spek
thin2 = Side(style="thin", color="BFBFBF")
bdr2 = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

outline_rows = [
    ("7", "OUTLINE SPEK", "11. OUTLINE_SPEK_Drive/", "Folder", "11 file dari Drive"),
    ("", "", "  ARS/2026.04.28 - OUTLINE SPEK ARSITEKTUR BIN - GEDUNG D.pdf", "PDF", "Arsitektur"),
    ("", "", "  ARS/2026.04.30 OUTLINE SPEK ARSITEKTUR BIN - GEDUNG D.pdf", "PDF", "Arsitektur terbaru"),
    ("", "", "  ARS/Archive/2026.04.24 - OUTLINE SPEK ARSITEKTUR GEDUNG D.pdf", "PDF", "Arsitektur lama"),
    ("", "", "  ARS/Archive/2026.04.24 - OUTLINE SPEK ARSITEKTUR GEDUNG D.xlsx", "XLSX", ""),
    ("", "", "  INT/OUTLINE SPEK INTERIOR BIN - GEDUNG D.pdf", "PDF", "Interior"),
    ("", "", "  INT/OUTLINE SPEK INTERIOR BIN_ kantor.pdf", "PDF", "Interior kantor"),
    ("", "", "  INF/RKS_Infrastruktur Gedung D.pdf", "PDF", "Infrastruktur"),
    ("", "", "  LAN/2026.03.08 OUTLINE SPEK LANSKAP - BIN.xlsx", "XLSX", "Lanskap"),
    ("", "", "  MEP/2026.04.21_OUTLINE SPEK ELEKTRIKAL GEDUNG KANTOR.pdf", "PDF", "Elektrikal"),
    ("", "", "  MEP/2026.04.21_OUTLINE SPEK MEKANIKAL GEDUNG KANTOR.pdf", "PDF", "Mekanikal"),
]

for i, (no, cat, name, ftype, ket) in enumerate(outline_rows, last_row + 1):
    ws_d.row_dimensions[i].height = 18
    
    cell_a = ws_d.cell(i, 1)
    cell_a.value = no; cell_a.alignment = center; cell_a.border = bdr2; cell_a.font = Font(size=9)
    
    cell_b = ws_d.cell(i, 2)
    if cat:
        cell_b.value = cat
        cell_b.fill = PatternFill("solid", fgColor="548235")
        cell_b.font = Font(color="FFFFFF", bold=True, size=9)
    cell_b.alignment = Alignment(horizontal="left", vertical="center")
    cell_b.border = bdr2
    
    cell_c = ws_d.cell(i, 3)
    if name.endswith("/"):
        cell_c.value = name; cell_c.font = Font(bold=True, size=9, color="2E75B6")
    elif name.startswith("  "):
        cell_c.value = name; cell_c.font = Font(size=8, color="444444")
    else:
        cell_c.value = name; cell_c.font = Font(size=9)
    cell_c.alignment = Alignment(horizontal="left", vertical="center")
    cell_c.border = bdr2
    
    cell_d = ws_d.cell(i, 4)
    cell_d.value = ftype; cell_d.alignment = center; cell_d.border = bdr2; cell_d.font = Font(size=9)
    
    cell_e = ws_d.cell(i, 5)
    cell_e.value = ket; cell_e.alignment = Alignment(horizontal="left", vertical="center")
    cell_e.border = bdr2; cell_e.font = Font(size=9)

wb.save(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")
print("✅ Updated: OUTLINE SPEK column added to SUMMARY + GEDUNG D sheet updated")
