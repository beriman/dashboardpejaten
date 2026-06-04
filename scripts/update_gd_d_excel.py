#!/usr/bin/env python3
"""Update sheet GEDUNG D di Excel berdasarkan data dari Google Drive spreadsheet."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.load_workbook(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")

# ── Styles ──
header_fill = PatternFill("solid", fgColor="375623")
header_font = Font(color="FFFFFF", bold=True, size=10)
sub_fill = PatternFill("solid", fgColor="548235")
sub_font = Font(color="FFFFFF", bold=True, size=9)
green_fill = PatternFill("solid", fgColor="C6EFCE")
green_font = Font(color="006100", size=9)
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
yellow_font = Font(color="9C5700", size=9)
red_fill = PatternFill("solid", fgColor="FFC7CE")
red_font = Font(color="9C0006", size=9)
na_fill = PatternFill("solid", fgColor="D9D9D9")
na_font = Font(color="7F7F7F", size=9, italic=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, val):
    cell.value = val
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

def sdr(cell, val):
    cell.value = val
    cell.fill = sub_fill
    cell.font = sub_font
    cell.alignment = left
    cell.border = border

def dat(cell, val, fill=None, font=None, align=left):
    cell.value = val
    if fill: cell.fill = fill
    if font: cell.font = font
    cell.alignment = align
    cell.border = border

# Hapus sheet GEDUNG D lama
if "GEDUNG D" in wb.sheetnames:
    del wb["GEDUNG D"]

ws = wb.create_sheet("GEDUNG D")
ws.sheet_properties.tabColor = "375623"
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 12

ws.merge_cells("A1:E1")
ws["A1"].value = "GEDUNG D — Detail File DED (Updated dari Google Drive)"
ws["A1"].font = Font(bold=True, size=12, color="375623")

h2 = ["NO", "KATEGORI", "NAMA FILE / FOLDER", "TIPE", "KET"]
for col, h in enumerate(h2, 1):
    hdr(ws.cell(2, col), h)

# Data berdasarkan Google Drive spreadsheet
# Sheet "Gedung D" menunjukkan:
# 1. Arsitektur - Submitted
# 2. Interior - Submitted  
# 3. Struktur - Submitted
# 4. Infrastruktur - Submitted
# 5. Mekanikal - Submitted
# 6. Elektrikal - Submitted
# 7. Outline Spek - Submitted
# 8. RKS - Submitted

d_data = [
    # ═══ ARSITEKTUR ═══
    ("1", "ARSITEKTUR", "01. ARSITEKTUR-20260504T035629Z-3-001/ (144 file)", "Folder", "Lengkap"),
    ("", "", "05. ARSITEKTUR_Drive/ (52 file dari Drive)", "Folder", "Revisi terbaru"),
    ("", "", "", "", ""),
    ("", "", "📁 00. EKSISTING/", "Folder", "Data existing 2018"),
    ("", "", "  - Gedung D EKSISTING - 10-01-2018.dwg", "DWG", ""),
    ("", "", "", "", ""),
    ("", "", "📁 01. CAD/", "Folder", "25+ subfolder"),
    ("", "", "  - 00. UMUM: COVER, SYMBOL, PERSPEKTIF, Masterplan", "Folder", ""),
    ("", "", "  - 01. SITE PLAN & BLOCK PLAN", "Folder", "AK0101-AK0102"),
    ("", "", "  - 02. DENAH General: LT1, LT2, LT3, LT4, ATAP", "Folder", "AK0201-AK0206"),
    ("", "", "  - 02. DENAH PARSIAL: Parsial LT1-ATAP (12 file)", "Folder", ""),
    ("", "", "  - 03. TAMPAK & POTONGAN: Tampak 1,2 Potong 1,2", "Folder", "AK0301-AK0304"),
    ("", "", "  - 04. PARSIAL FASAD: Selatan, Timur, Utara, Barat", "Folder", "AK0401-AK0407"),
    ("", "", "  - 05. DETAIL FASAD: Perforated, Gawangan, Jendela", "Folder", "AK0501-AK0503"),
    ("", "", "  - 06. DETAIL CORE: Denah, Pola Lantai, Pola Plafon", "Folder", "AK0601"),
    ("", "", "  - 07. DETAIL TANGGA: AK0701-AK0702 (8 file)", "Folder", ""),
    ("", "", "  - 08. DETAIL TOILET: Denah, Potongan, Legend", "Folder", "AK0801-AK0802"),
    ("", "", "  - 09. POLA LANTAI: LT1-ATAP (15 file DWG+PDF)", "Folder", "AK0901-AK0905"),
    ("", "", "  - 10. POLA PLAFON: LT1-ATAP (15 file)", "Folder", "AK1001-AK1005"),
    ("", "", "  - 11. SKEMA KUSEN: AK1101-AK1108 (8 file)", "Folder", ""),
    ("", "", "  - 12. DETAIL KHUSUS", "Folder", ""),
    ("", "", "  - 13. DETAIL MEP", "Folder", ""),
    ("", "", "  - 14. SARANA LUAR: Drop Off, Planter", "Folder", "AK1401"),
    ("", "", "", "", ""),
    ("", "", "📁 02. PDF/", "Folder", "PDF consolidated"),
    ("", "", "📁 03. RKS & OUTLINE SPEK/", "Folder", "RKS Arsitektur"),
    ("", "", "  - Ars RKS & Outline Spek", "Folder", ""),
    # ═══ INTERIOR ═══
    ("2", "INTERIOR", "02. INTERIOR/ (41 file)", "Folder", "Lengkap"),
    ("", "", "06. INTERIOR_Drive/ (109 file dari Drive)", "Folder", "Lebih lengkap"),
    ("", "", "", "", ""),
    ("", "", "📁 01. CAD/", "Folder", "18 file"),
    ("", "", "  - INT 0: COVER, Scope Interior, Drawing List", "", ""),
    ("", "", "  - INT 1: DENAH GENERAL", "Folder", ""),
    ("", "", "  - INT 2: DETAIL RUANGAN (14 ruangan)", "", ""),
    ("", "", "    LOBBY, STAFF LT3, DIREKTUR, DEPUTI, ESELON 3", "", ""),
    ("", "", "    RAPAT DIREKTORAT, STAFF PERCETAKAN, AHLI MADYA", "", ""),
    ("", "", "    TUNGGU, TELECONFERENCE, RAPAT BESAR", "", ""),
    ("", "", "    STAFF DIREKTORAT, AHLI MADYA LT3-4", "", ""),
    ("", "", "", "", ""),
    ("", "", "📁 02. PDF/", "Folder", "18 file PDF"),
    ("", "", "  - RKS INTERIOR BIN - GEDUNG D.pdf", "PDF", ""),
    ("", "", "  - BQ Interior gedung D.xlsx", "XLSX", ""),
    # ═══ STRUKTUR ═══
    ("3", "STRUKTUR", "03. STRUKTUR_Drive/ (19 file dari Drive)", "Folder", "Termasuk X_KOP"),
    ("", "", "  - X_KOP_Gedung D_Rev1.dwg", "DWG", "Kop terbaru"),
    # ═══ INFRASTRUKTUR ═══
    ("4", "INFRASTRUKTUR", "08. INFRASTRUKTUR_Drive/ (1 file)", "Folder", ""),
    # ═══ MEKANIKAL ═══
    ("5", "MEKANIKAL", "05. MEKANIKAL-20260504T035631Z-3-001/ (74 file)", "Folder", "Lengkap"),
    ("", "", "(09. MEKANIKAL_Drive/ kosong)", "", ""),
    ("", "", "", "", ""),
    ("", "", "📁 01. AC/", "Folder", "DWG"),
    ("", "", "📁 02. LIFT/", "Folder", "DWG"),
    ("", "", "📁 03. PLUMBING/", "Folder", "DWG"),
    ("", "", "📁 04. HYDRANT/", "Folder", "DWG"),
    ("", "", "📁 05. SPRINKLER/", "Folder", "DWG"),
    ("", "", "  - DED MEKANIKAL GEDUNG KANTOR D.pdf", "PDF", "Konsolidated"),
    ("", "", "  - RKS & OUTLINE SPEK MEKANIKAL GEDUNG KANTOR.pdf", "PDF", ""),
    # ═══ ELEKTRIKAL ═══
    ("6", "ELEKTRIKAL", "06. ELEKTRIKAL/ (86 file)", "Folder", "Lengkap"),
    ("", "", "10. ELEKTRIKAL_Drive/ (86 file dari Drive)", "Folder", "Mirror"),
    ("", "", "", "", ""),
    ("", "", "📁 LAK (Elektrikal Daya):", "Folder", "9 subfolder"),
    ("", "", "  - 00. DAFTAR GAMBAR, 01. SITEPLAN, 02. SYSTEM", "", ""),
    ("", "", "  - 03. PENCAHAYAAN, 04. KOTAK KONTAK", "", ""),
    ("", "", "  - 05. SCHEDULE PANEL, 06. R. GENSET", "", ""),
    ("", "", "  - 07. PEREDAM PETIR, 08. KABEL TRAY, 09. DETAIL", "", ""),
    ("", "", "", "", ""),
    ("", "", "📁 LAL (Low Voltage):", "Folder", "5 subfolder"),
    ("", "", "  - 01. DAFTAR GAMBAR, 02. FIRE ALARM", "", ""),
    ("", "", "  - 03. DATA/TELPON/IPTV, 04. SECURITY", "", ""),
    ("", "", "  - 05. CONFERANCE ROOM SYSTEM", "", ""),
    ("", "", "", "", ""),
    ("", "", "  - DED ELEKTRIKAL GEDUNG KANTOR D.pdf", "PDF", "Konsolidated"),
    ("", "", "  - RKS & OUTLINE SPEK ELEKTRIKAL GEDUNG KANTOR.pdf", "PDF", ""),
    # ═══ OUTLINE SPEK ═══
    ("7", "OUTLINE SPEK", "11. OUTLINE_SPEK_Drive/ (11 file dari Drive)", "Folder", "Lebih banyak"),
    # ═══ RKS ═══
    ("8", "RKS", "12. RKS_Drive/ (1 file dari Drive)", "Folder", ""),
    # ═══ RAB ═══
    ("9", "RAB", "2026.05.22 RAB GD. KANTOR BIN (Konstruksi).xlsx", "XLSX", "JAKON"),
    ("", "", "2026.05.22 RAB GD. KANTOR BIN (Pengadaan).xlsx", "XLSX", "JAKON"),
    # ═══ LAINNYA ═══
    ("10", "LAINNYA", "2026.05.11_Gedung D_Revisi GWT.dwg (07. STRUKTUR_Drive)", "DWG", "Revisi GWT"),
    ("", "", "STP Biofilter.dwg (07. STRUKTUR_Drive)", "DWG", ""),
]

for i, row_data in enumerate(d_data, 3):
    ws.row_dimensions[i].height = 18
    for j, val in enumerate(row_data, 1):
        cell = ws.cell(i, j)
        if j == 2 and val and not val.startswith("  ") and not val.startswith("📁"):
            sdr(cell, val)
        elif val.startswith("📁"):
            cell.value = val
            cell.font = Font(bold=True, size=9, color="2E75B6")
            cell.alignment = left
        elif val.startswith("  -"):
            cell.value = val
            cell.font = Font(size=8, color="444444")
            cell.alignment = left
        elif val:
            cell.value = val
            cell.font = Font(size=9)
            cell.alignment = left if j != 1 else center
        cell.border = border

ws.freeze_panes = "A3"

# ═══ Update SUMMARY sheet untuk Gedung D ═══
ws_sum = wb["SUMMARY"]
# Update kolom Gedung D (kolom C) di SUMMARY
# Baris 5 = header, data mulai baris 6
# Cari baris yang sesuai dan update

# Update row headers untuk Gedung D di SUMMARY
sum_updates = {
    # Arsitektur
    7: "✅ ADA (144+52 file)",   # Denah
    8: "✅ ADA",   # Tampak
    9: "✅ ADA",   # Pola Lantai
    10: "✅ ADA",  # Plafond
    11: "✅ ADA",  # Kusen
    12: "✅ ADA",  # Toilet
    13: "✅ ADA",  # Tangga
    14: "✅ ADA",  # Fasad
    15: "✅ ADA",  # Schedule Material
    16: "✅ ADA",  # Ars PDF
    17: "~65 DWG", # DWG count
    # Struktur
    19: "✅ ADA",
    20: "✅ ADA",
    21: "✅ ADA",
    22: "✅ ADA",
    23: "✅ ADA",
    24: "✅ ADA (SD-01~06)",
    25: "✅ ADA (di Struktur Drive)",
    26: "~20 DWG",
    # MEP
    28: "✅ ADA (lengkap)",
    29: "✅ ADA",
    30: "✅ ADA",
    31: "✅ ADA (Lengkap, 80+ DWG)",
    32: "✅ ADA",
    33: "LIFT: ✅ di MEP Drive",
    34: "✅ ADA",
    35: "~80+ DWG",
    # Interior
    37: "", 38: "", 39: "", 40: "",
    41: "✅ ADA (41+109 file)",
    42: "✅ ADA",
    # Site Dev
    44: "✅ ADA",
    45: "✅ ADA",
    # RKS
    47: "", 48: "",
    49: "✅ ADA (Drive)",  # RKS Interior
    50: "✅ ADA (Drive)",  # RKS Elektrikal
    # RAB
    52: "", 53: "",
    54: "✅ ADA",
    # Laporan
    56: "❌", 57: "❌", 58: "❌", 59: "❌",
    # Outline Spek
    61: "✅ ADA (Drive: 11 file)",
}

for row_num, val in sum_updates.items():
    cell = ws_sum.cell(row_num, 3)  # Column C = Gedung D
    if val:
        cell.value = val
        if "✅" in val:
            cell.fill = green_fill
            cell.font = green_font
            cell.alignment = center
        elif "❌" in val:
            cell.fill = na_fill
            cell.font = na_font
            cell.alignment = center
        else:
            cell.fill = yellow_fill
            cell.font = yellow_font
            cell.alignment = center
        cell.border = border

wb.save(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")
print("✅ Excel updated")
