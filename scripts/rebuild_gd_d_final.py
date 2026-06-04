#!/usr/bin/env python3
"""
Rebuild GEDUNG D sheet dari data folder yang discan dari Google Drive.
Struktur sejajar dengan GEDUNG B dan GEDUNG K.
Kolom: NO | KATEGORI | NAMA FILE / FOLDER | TIPE | KET
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.load_workbook(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")

# ── Styles (sama dengan GEDUNG B) ──
hdr_fill = PatternFill("solid", fgColor="375623")
hdr_font = Font(color="FFFFFF", bold=True, size=10)
sub_fill = PatternFill("solid", fgColor="548235")
sub_font = Font(color="FFFFFF", bold=True, size=9)
thin = Side(style="thin", color="BFBFBF")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
left = Alignment(horizontal="left", vertical="center")
ctr  = Alignment(horizontal="center", vertical="center")

def hdr(cell, val):
    cell.value = val; cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = ctr; cell.border = bdr

def cat(cell, val):
    cell.value = val; cell.fill = sub_fill; cell.font = sub_font
    cell.alignment = left; cell.border = bdr

def dat(cell, val, bold=False, color="000000", size=9, align=left):
    cell.value = val
    cell.font = Font(bold=bold, size=size, color=color)
    cell.alignment = align; cell.border = bdr

# Delete old GEDUNG D
del wb["GEDUNG D"]
ws = wb.create_sheet("GEDUNG D")
ws.sheet_properties.tabColor = "375623"
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 48
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 14

ws.merge_cells("A1:E1")
ws["A1"].value = "GEDUNG D — Detail File DED"
ws["A1"].font = Font(bold=True, size=12, color="375623")

for c, h in enumerate(["NO", "KATEGORI", "NAMA FILE / FOLDER", "TIPE", "KET"], 1):
    hdr(ws.cell(2, c), h)

# ═══ DATA — setiap baris punya NO dan KATEGORI (sejajar) ═══
data = [
    # ── ARSITEKTUR ──
    ("1", "ARSITEKTUR", "01. ARSITEKTUR-20260504T035629Z-3-001/", "Folder", "144 file lengkap"),
    ("", "", "01. CAD/00. UMUM — COVER, SYMBOL, PERSPEKTIF, Masterplan", "Folder", ""),
    ("", "", "01. CAD/01. SITE PLAN & BLOCK PLAN — AK0101-AK0102", "Folder", ""),
    ("", "", "01. CAD/02. DENAH GENERAL — LT1, LT2, LT3, LT4, ATAP, ATAP-2", "Folder", "6 DWG"),
    ("", "", "01. CAD/02. DENAH PARSIAL — Parsial LT1 s.d. ATAP", "Folder", "12 DWG"),
    ("", "", "01. CAD/03. TAMPAK & POTONGAN — Tampak 1,2 + Potong 1,2", "Folder", "4 DWG + 2 PCP"),
    ("", "", "01. CAD/04. PARSIAL FASAD — Selatan, Timur, Utara, Barat", "Folder", "14 DWG"),
    ("", "", "01. CAD/05. DETAIL FASAD — Perforated, Gawangan, Jendela", "Folder", "5 DWG"),
    ("", "", "01. CAD/06. DETAIL CORE — Denah, Pola Lantai, Pola Plafon, Potong", "Folder", "4 DWG"),
    ("", "", "01. CAD/07. DETAIL TANGGA — AK0701-AK0702", "Folder", "8 DWG"),
    ("", "", "01. CAD/08. DETAIL TOILET — Denah, Potongan, Legend", "Folder", "10+ DWG + RAR"),
    ("", "", "01. CAD/09. POLA LANTAI — LT1 s.d. ATAP Parsial", "Folder", "15 DWG + PDF"),
    ("", "", "01. CAD/10. POLA PLAFON — LT1 s.d. ATAP Parsial", "Folder", "15 DWG"),
    ("", "", "01. CAD/11. SKEMA KUSEN — AK1101-AK1108", "Folder", "8 DWG"),
    ("", "", "01. CAD/12. DETAIL KHUSUS", "Folder", ""),
    ("", "", "01. CAD/13. DETAIL MEP", "Folder", ""),
    ("", "", "01. CAD/14. SARANA LUAR — Drop Off, Planter", "Folder", ""),
    ("", "", "01. CAD/15. DETAIL STANDAR", "Folder", ""),
    ("", "", "02. PDF/", "Folder", ""),
    ("", "", "03. RKS & OUTLINE SPEK/", "Folder", "RKS Arsitektur"),
    ("", "", "05. ARSITEKTUR_Drive/ — 52 file revisi terbaru", "Folder", "Dari Google Drive"),
    # ── INTERIOR ──
    ("2", "INTERIOR", "02. INTERIOR/", "Folder", "41 file lengkap"),
    ("", "", "01. CAD/INT 0 — COVER, Scope Interior, Drawing List", "Folder", ""),
    ("", "", "01. CAD/INT 1 — DENAH GENERAL", "Folder", ""),
    ("", "", "01. CAD/INT 2 — DETAIL RUANGAN (14 ruangan)", "Folder", ""),
    ("", "", "  R: LOBBY, STAFF LT3, DIREKTUR, DEPUTI, ESELON 3", "", ""),
    ("", "", "  R: RAPAT DIREKTORAT, STAFF PERCETAKAN, AHLI MADYA", "", ""),
    ("", "", "  R: TUNGGU, TELECONFERENCE, RAPAT BESAR, STAFF DIREKTORAT", "", ""),
    ("", "", "02. PDF/ — 18 file PDF", "Folder", ""),
    ("", "", "03. RKS & OUTLINE SPEK/RKS INTERIOR BIN - GEDUNG D.pdf", "PDF", ""),
    ("", "", "BQ Interior gedung D.xlsx", "XLSX", ""),
    ("", "", "06. INTERIOR_Drive/ — 109 file dari Drive (lebih lengkap)", "Folder", ""),
    # ── STRUKTUR ──
    ("3", "STRUKTUR", "03. STRUKTUR-20260504T035630Z-3-001/", "Folder", "CAD + PDF + RKS"),
    ("", "", "07. STRUKTUR_Drive/ — 19 file dari Drive", "Folder", "Termasuk X_KOP"),
    # ── INFRASTRUKTUR ──
    ("4", "INFRASTRUKTUR", "04. INFRASTRUKTUR/", "Folder", "CAD + PDF + RKS"),
    ("", "", "04. INFRASTRUKTUR-20260504T035630Z-3-001/", "Folder", "Mirror dari Drive"),
    ("", "", "08. INFRASTRUKTUR_Drive/", "Folder", "1 file"),
    # ── MEKANIKAL ──
    ("5", "MEKANIKAL", "05. MEKANIKAL-20260504T035631Z-3-001/", "Folder", "74 file lengkap"),
    ("", "", "01. CAD/01. AC/", "Folder", "DWG"),
    ("", "", "01. CAD/02. LIFT/", "Folder", "DWG"),
    ("", "", "01. CAD/03. PLUMBING/", "Folder", "DWG"),
    ("", "", "01. CAD/04. HYDRANT/", "Folder", "DWG"),
    ("", "", "01. CAD/05. SPRINKLER/", "Folder", "DWG"),
    ("", "", "02. PDF/DED MEKANIKAL GEDUNG KANTOR D.pdf", "PDF", "Konsolidated"),
    ("", "", "03. RKS & OUTLINE SPEK/RKS & OUTLINE SPEK MEKANIKAL.pdf", "PDF", ""),
    # ── ELEKTRIKAL ──
    ("6", "ELEKTRIKAL", "06. ELEKTRIKAL/", "Folder", "86 file lengkap"),
    ("", "", "06. ELEKTRIKAL-20260504T035637Z-3-001/", "Folder", "Mirror dari Drive"),
    ("", "", "10. ELEKTRIKAL_Drive/ — 86 file dari Drive", "Folder", ""),
    ("", "", "01. CAD/LAK/ (9 subsistem daya)", "Folder", "Siteplan, System, Pencahayaan, Kotak Kontak, Panel, Genset, Petir, Kabel Tray, Detail"),
    ("", "", "01. CAD/LAL/ (5 subsistem low voltage)", "Folder", "Fire Alarm, Data/Telpon, Security, Conference, Daftar Gambar"),
    ("", "", "02. PDF/DED ELEKTRIKAL GEDUNG KANTOR D.pdf", "PDF", "Konsolidated"),
    ("", "", "03. RKS & OUTLINE SPEK/RKS & OUTLINE SPEK ELEKTRIKAL.pdf", "PDF", ""),
    # ── OUTLINE SPEK ──
    ("7", "OUTLINE SPEK", "11. OUTLINE_SPEK_Drive/", "Folder", "11 file dari Google Drive"),
    ("", "", "  ARS/2026.04.28 — OUTLINE SPEK ARSITEKTUR BIN GEDUNG D.pdf", "PDF", ""),
    ("", "", "  ARS/2026.04.30 — OUTLINE SPEK ARSITEKTUR BIN GEDUNG D.pdf", "PDF", "Terbaru"),
    ("", "", "  ARS/Archive/2026.04.24 — OUTLINE SPEK ARSITEKTUR GEDUNG D (.pdf + .xlsx)", "PDF/XLSX", ""),
    ("", "", "  INT/ — OUTLINE SPEK INTERIOR BIN GEDUNG D (2 file)", "PDF", ""),
    ("", "", "  INF/ — RKS Infrastruktur Gedung D.pdf", "PDF", ""),
    ("", "", "  LAN/ — OUTLINE SPEK LANSKAP BIN.xlsx", "XLSX", ""),
    ("", "", "  MEP/ — OUTLINE SPEK ELEKTRIKAL GEDUNG KANTOR.pdf", "PDF", ""),
    ("", "", "  MEP/ — OUTLINE SPEK MEKANIKAL GEDUNG KANTOR.pdf", "PDF", ""),
    # ── RKS ──
    ("8", "RKS", "01. ARSITEKTUR/03. RKS & OUTLINE SPEK/", "Folder", "RKS Arsitektur di folder lama"),
    ("", "", "02. INTERIOR/03. RKS & OUTLINE SPEK/RKS INTERIOR BIN - GEDUNG D.pdf", "PDF", ""),
    ("", "", "03. STRUKTUR/03. RKS & OUTLINE SPEK/RKS-Str-BIN-Kantor.pdf", "PDF", ""),
    ("", "", "04. INFRASTRUKTUR/03. RKS & OUTLINE SPEK/RKS_Infrastruktur Gedung D.pdf", "PDF", ""),
    ("", "", "05. MEKANIKAL/03. RKS & OUTLINE SPEK/RKS & OUTLINE SPEK MEKANIKAL.pdf", "PDF", ""),
    ("", "", "06. ELEKTRIKAL/03. RKS & OUTLINE SPEK/RKS & OUTLINE SPEK ELEKTRIKAL.pdf", "PDF", ""),
    ("", "", "12. RKS_Drive/ — 1 file dari Drive", "Folder", ""),
    # ── RAB ──
    ("9", "RAB", "2026.05.22 RAB GD. KANTOR BIN (Konstruksi) - konsinyering JAKON.xlsx", "XLSX", ""),
    ("", "", "2026.05.22 RAB GD. KANTOR BIN (Pengadaan) - Konsinyering JAKON.xlsx", "XLSX", ""),
    # ── LAINNYA ──
    ("10", "LAINNYA", "2026.05.11_Gedung D_Revisi GWT.dwg", "DWG", "Revisi GWT"),
    ("", "", "STP Biofilter.dwg", "DWG", ""),
]

current_cat = None
for i, row_data in enumerate(data, 3):
    ws.row_dimensions[i].height = 18
    no, category, name, ftype, ket = row_data
    
    # NO
    cell_a = ws.cell(i, 1)
    cell_a.value = no; cell_a.alignment = ctr; cell_a.border = bdr; cell_a.font = Font(size=9)
    
    # KATEGORI
    cell_b = ws.cell(i, 2)
    if category and category != current_cat:
        cat(cell_b, category); current_cat = category
    else:
        cell_b.value = None; cell_b.border = bdr
    
    # NAMA FILE
    cell_c = ws.cell(i, 3)
    if name.endswith("/"):
        dat(cell_c, name, bold=True, color="2E75B6")
    elif name.startswith("  R:"):
        dat(cell_c, name, size=8, color="666666")
    elif name.startswith("  "):
        dat(cell_c, name, size=8, color="444444")
    else:
        dat(cell_c, name)
    
    # TIPE
    cell_d = ws.cell(i, 4)
    cell_d.value = ftype; cell_d.alignment = ctr; cell_d.border = bdr; cell_d.font = Font(size=9)
    
    # KET
    cell_e = ws.cell(i, 5)
    cell_e.value = ket; cell_e.alignment = left; cell_e.border = bdr; cell_e.font = Font(size=9)

ws.freeze_panes = "A3"

# ═══ Update SUMMARY — kolom OUTLINE SPEK ═══
ws_sum = wb["SUMMARY"]
green_fill = PatternFill("solid", fgColor="C6EFCE")
green_font = Font(color="006100", size=9, bold=True)
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
yellow_font = Font(color="9C5700", size=9)
na_fill2 = PatternFill("solid", fgColor="F2F2F2")
na_font2 = Font(color="7F7F7F", size=9, italic=True)

outline_status = {
    "ARSITEKTUR": "ADA",
    "STRUKTUR": "BELUM",
    "MEP (MEKANIKAL ELEKTRONIK)": "BELUM",
    "INTERIOR": "BELUM",
    "SITE DEVELOPMENT": "BELUM",
    "OUTLINE SPESIFIKASI": "ADA",
}

for row_num in range(4, ws_sum.max_row + 1):
    disc = ws_sum.cell(row_num, 1).value
    cell = ws_sum.cell(row_num, 5)
    if disc and not str(disc).startswith("  "):
        s = outline_status.get(str(disc).strip(), "—")
        if s == "ADA":
            cell.value = "ADA"; cell.fill = green_fill; cell.font = green_font
        elif s == "BELUM":
            cell.value = "BELUM"; cell.fill = yellow_fill; cell.font = yellow_font
        else:
            cell.value = "—"; cell.fill = na_fill2; cell.font = na_font2
    else:
        cell.value = "—"; cell.fill = na_fill2; cell.font = na_font2
    cell.alignment = ctr; cell.border = bdr

wb.save(r"H:\My Drive\Work in Progress\02 DED (gambar dari Perencana)\DED_Checklist_Per_Gedung.xlsx")
print("✅ GEDUNG D rebuilt from scratch — format sejajar B/K, OUTLINE SPEK di posisi benar")
print("✅ SUMMARY kolom OUTLINE SPEK diisi")
