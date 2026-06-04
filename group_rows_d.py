#!/usr/bin/env python3
"""Add row grouping (outline) to Gedung D sheet for collapsible sections."""

import openpyxl
import shutil
import os

EXCEL_PATH = r"H:\My Drive\Work in Progress\06 Quantity Take Off\Bekisting\Perhitungan_Bekisting_Pejaten.xlsx"
TEMP_PATH = r"C:\Users\bim\dashboardpejaten\grouped_temp.xlsx"

# Copy to temp first
shutil.copy2(EXCEL_PATH, TEMP_PATH)

wb = openpyxl.load_workbook(TEMP_PATH)
ws = wb['Gedung D']

# Find section ranges
sections = []
current_section = None
header_row = None

for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
    val = row[0]
    if val and isinstance(val, str) and 'elemen' in val.lower():
        current_section = val
        header_row = i
    elif val and isinstance(val, str) and val.startswith('Total ') and current_section:
        start_row = header_row + 1
        end_row = i - 1
        if start_row <= end_row:
            sections.append({
                'name': current_section,
                'start': start_row,
                'end': end_row,
            })
        current_section = None

print(f"Found {len(sections)} sections:")
for s in sections:
    print(f"  {s['name']}: rows {s['start']}-{s['end']}")

# Add outline levels
for s in sections:
    for row_idx in range(s['start'], s['end'] + 1):
        ws.row_dimensions[row_idx].outline_level = 1
        ws.row_dimensions[row_idx].hidden = False

ws.sheet_properties.outlinePr.summaryBelow = True

wb.save(TEMP_PATH)
print(f"\nSaved to temp: {TEMP_PATH}")

# Now try to copy back
try:
    shutil.copy2(TEMP_PATH, EXCEL_PATH)
    print(f"✅ Copied back to: {EXCEL_PATH}")
except PermissionError:
    print(f"⚠️  Could not copy back (file may be open in Excel)")
    print(f"   Temp file saved at: {TEMP_PATH}")
    print(f"   Please close Excel and copy manually, or run this script again.")
