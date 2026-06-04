#!/usr/bin/env python3
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(r"H:\My Drive\Work in Progress\03 Shop Drawing\Monitoring_SD_Pejaten.xlsx", data_only=True)
ws = wb["Gedung K"]

print("=== Gedung K - Target Pengumpulan SD ===\n")
count = 0
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
    judul = row[1]
    target = row[3]
    if judul and target and isinstance(target, datetime):
        print(f"  R{i}: {str(julul)[:45]:<45} -> {target.strftime('%d %b %Y')}")
        count += 1
    elif judul and not target:
        print(f"  R{i}: {str(judul)[:45]:<45} -> [NO TARGET]")

print(f"\nTotal: {count} SD with target dates")
