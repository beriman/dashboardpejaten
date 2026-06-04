#!/usr/bin/env python3
"""
Update target pengumpulan Shop Drawing Gedung K
berdasarkan jadul konstruksi (Gantt chart) yang dikirim user.
Target = tanggal mulai pekerjaan - 4 hari kerja.
"""

import openpyxl, re
from datetime import datetime, timedelta
from copy import copy

EXCEL_PATH = r"H:\My Drive\Work in Progress\03 Shop Drawing\Monitoring_SD_Pejaten.xlsx"

# ─── Jadwal konstruksi Gedung K (dari Gantt chart, bottom-up) ───
# Format: (tanggal_mulai, pekerjaan, keterangan)
# Target SD = tanggal_mulai - 4 hari kerja

def workday_before(start_date, days=4):
    """Hitung tanggal 'days' hari kerja sebelum start_date (Sabtu-Minggu tidak dihitung)."""
    d = start_date
    count = 0
    while count < days:
        d -= timedelta(days=1)
        # 0=Senin, 5=Sabtu, 6=Minggu
        if d.weekday() < 5:  # Senin-Jumat
            count += 1
    return d

# Jadwal dari Gantt chart (perkiraan berdasarkan gambar)
# Borepile: ~7 Mei - 10 Juni
# PC + TB: ~11 Juni - 20 Juni
# LT 1: ~11 Juni - 20 Juni
# LT 2: ~15 Juni - 1 Juli
# LT 3: ~25 Juni - 9 Juli
# LT 4: ~9 Juli - 23 Juli
# LT 5: ~17 Juli - 14 Agustus
# ATAP: ~8 Agustus - 5 September

schedule = {
    'BOREPILE': {
        'start': datetime(2026, 5, 7),
        'items': [
            'STANDAR DETAIL 1', 'STANDAR DETAIL 2', 'STANDAR DETAIL 3',
            'STANDAR DETAIL 4', 'STANDAR DETAIL 5',
            'DENAH TIANG PANCANG',
            'NILAI KOORDINAT TITIK PANCANG GEDUNG K',
            'DETAIL TIANG PANCANG',
        ]
    },
    'PC_TB': {
        'start': datetime(2026, 6, 11),
        'items': [
            'DENAH PILE CAP',
            'DETAIL PILE CAP',
            'DENAH TIE BEAM',
            'DENAH BALOK LANTAI 1 / TIE BEAM',
        ]
    },
    'LT_1': {
        'start': datetime(2026, 6, 11),
        'items': [
            'DENAH KOLOM',
            'DENAH BALOK LANTAI 1',
            'TABEL BALOK',
            'TABEL PENULANGAN BALOK GEDUNG K',
            'DENAH STRUKTUR GWT & R. POMPA',
        ]
    },
    'LT_2': {
        'start': datetime(2026, 6, 15),
        'items': [
            'DENAH BALOK LANTAI 2',
        ]
    },
    'LT_3': {
        'start': datetime(2026, 6, 25),
        'items': [
            'DENAH BALOK LANTAI 3',
            'DETAIL PENULANGAN SHEAR WALL',
        ]
    },
    'LT_4': {
        'start': datetime(2026, 7, 9),
        'items': [
            'DENAH BALOK LANTAI 4',
        ]
    },
    'LT_5': {
        'start': datetime(2026, 7, 17),
        'items': [
            'DENAH BALOK LANTAI 5',
        ]
    },
    'ATAP': {
        'start': datetime(2026, 8, 8),
        'items': [
            'DENAH BALOK LANTAI DAK',
            'DENAH BALOK ATAP',
        ]
    },
    'ARSITEKTUR': {
        'start': datetime(2026, 8, 7),
        'items': [
            'POTONGAN A-B', 'SITE PLAN', 'BLOCK PLAN',
            'DENAH LANTAI 1', 'DENAH LANTAI 2', 'DENAH LANTAI 3',
            'DENAH LANTAI 4', 'DENAH LANTAI 5', 'DENAH DAK ATAP', 'DENAH ATAP',
            'DENAH DINDING GEDUNG SERVER GEDUNG K',
            'DETAIL TOILET', 'TAMPAK DEPAN', 'TAMPAK BELAKANG', 'TAMPAK SAMPING',
            'DENAH POLA LANTAI 1', 'DENAH POLA LANTAI 2', 'DENAH POLA LANTAI 3',
            'DENAH POLA LANTAI 4', 'DENAH POLA LANTAI 5', 'DENAH POLA LANTAI DAK',
            'DENAH PLAFON LANTAI 1', 'DENAH PLAFON LANTAI 2', 'DENAH PLAFON LANTAI 3',
            'DENAH PLAFON LANTAI 4', 'DENAH PLAFON LANTAI 5',
            'DENAH TITIK LAMPU LT. 1', 'DENAH TITIK LAMPU LT. 2',
            'DENAH TITIK LAMPU LT. 3', 'DENAH TITIK LAMPU LT. 4',
            'DENAH TITIK LAMPU LT. 5',
            'DENAH TANGGA', 'DETAIL TANGGA',
            'DENAH RENCANA KUSEN LT. 1', 'DENAH RENCANA KUSEN LT. 2',
            'DENAH RENCANA KUSEN LT. 5', 'DENAH RENCANA KUSEN LT. DAK ATAP',
            'DETAIL KUSEN #1', 'DETAIL KUSEN #2', 'DETAIL KUSEN #4',
            'DETAIL KUSEN #5', 'DETAIL KUSEN #6', 'DETAIL KUSEN #7',
            'DETAIL KUSEN #8', 'DETAIL KUSEN #9',
            'SITE MANAGEMENT GEDUNG K',
        ]
    },
    'INTERIOR': {
        'start': datetime(2026, 8, 7),
        'items': [
            'DENAH INTERIOR LT.1', 'DENAH INTERIOR LT.2', 'DENAH INTERIOR LT.3',
            'DENAH INTERIOR LT.4', 'DENAH INTERIOR LT.5',
        ]
    },
    'MEP': {
        'start': datetime(2026, 7, 15),
        'items': [
            'DENAH INSTALASI TELEPON LT.1', 'DENAH INSTALASI TELEPON LT.2',
            'DENAH INSTALASI TELEPON LT.3', 'DENAH INSTALASI TELEPON LT.4',
            'DENAH INSTALASI DATA LT.1', 'DENAH INSTALASI DATA LT.2',
            'DENAH INSTALASI DATA LT.3', 'DENAH INSTALASI DATA LT.4',
            'DENAH INSTALASI TATA SUARA LT.1', 'DENAH INSTALASI TATA SUARA LT.2',
            'DENAH INSTALASI TATA SUARA LT.3', 'DENAH INSTALASI TATA SUARA LT.4',
            'DIAGRAM SKEMATIK PRESSURIZED',
            'DIAGRAM SKEMATIK EXHAUST LT.1', 'DIAGRAM SKEMATIK EXHAUST LT.2',
            'DIAGRAM SKEMATIK EXHAUST LT.3', 'DIAGRAM SKEMATIK EXHAUST LT.4',
        ]
    },
}

# Build mapping: judul_gambar → target_date
target_map = {}
for phase, data in schedule.items():
    target = workday_before(data['start'], 4)
    for item in data['items']:
        target_map[item.upper()] = target

print("Target mapping built:")
for phase, data in schedule.items():
    target = workday_before(data['start'], 4)
    print(f"  {phase}: start={data['start'].strftime('%d %b %Y')} → target={target.strftime('%d %b %Y')}")

# ─── Load Excel and update ───
print(f"\nLoading Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['Gedung K']

# Find the column indices
# Row 4: ('No', 'Judul Gambar', 'Nomor Gambar', 'Target Pengumpulan', 'Realisasi ke MK', 'Approved MK', 'Distribusi')
judul_col = 2  # Column B
target_col = 4  # Column D

updated = 0
skipped = 0
not_found = []

# Build a list of all SD titles in the target_map for matching
target_titles = set(target_map.keys())

# Iterate through all rows
for row_idx in range(5, ws.max_row + 1):
    judul = ws.cell(row_idx, judul_col).value
    if not judul or not isinstance(judul, str):
        continue
    
    judul_clean = judul.strip().upper()
    
    # Try exact match first
    if judul_clean in target_map:
        target_date = target_map[judul_clean]
        ws.cell(row_idx, target_col, target_date)
        updated += 1
        continue
    
    # Try partial match (for items with ↳ prefix or slightly different naming)
    matched = False
    for target_title, target_date in target_map.items():
        # Check if the target title is contained in the judul or vice versa
        if target_title in judul_clean or judul_clean in target_title:
            ws.cell(row_idx, target_col, target_date)
            updated += 1
            matched = True
            break
    
    if not matched:
        # Check if it's a sub-item (starts with spaces/↳)
        # Try to match with the parent item
        judul_normalized = re.sub(r'^[↳\s]+', '', judul_clean).strip()
        if judul_normalized in target_map:
            target_date = target_map[judul_normalized]
            ws.cell(row_idx, target_col, target_date)
            updated += 1
            matched = True
        
        if not matched:
            # Check partial match on normalized title
            for target_title, target_date in target_map.items():
                if target_title in judul_normalized or judul_normalized in target_title:
                    ws.cell(row_idx, target_col, target_date)
                    updated += 1
                    matched = True
                    break
    
    if not matched:
        not_found.append((row_idx, judul))

print(f"\nUpdated: {updated} rows")
print(f"Not matched: {len(not_found)} rows")

if not_found:
    print("\nUnmatched items:")
    for row_idx, judul in not_found:
        print(f"  R{row_idx}: {judul}")

# ─── Save ───
wb.save(EXCEL_PATH)
print(f"\n✅ Saved: {EXCEL_PATH}")

# ─── Verify ───
print("\n=== Verification ===")
wb2 = openpyxl.load_workbook(EXCEL_PATH)
ws2 = wb2['Gedung K']
for row_idx in range(5, min(30, ws2.max_row + 1)):
    judul = ws2.cell(row_idx, judul_col).value
    target = ws2.cell(row_idx, target_col).value
    if judul and target:
        print(f"  R{row_idx}: {str(judul)[:40]:<40} → {target}")
