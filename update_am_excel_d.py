#!/usr/bin/env python3
"""
Update Monitoring Approval Material Gedung D berdasarkan Outline Spec terbaru.
File: Monitoring Approval Material Gedung D.xlsx
"""

import openpyxl
from copy import copy
from datetime import datetime

EXCEL_PATH = r"H:\My Drive\Work in Progress\07 Quality Control\Approval Material\Monitoring Approval Material Gedung D.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Monitoring Gedung D"]

# Column mapping (1-indexed):
# 1=No, 2=ID, 3=Gedung, 4=Disiplin, 5=Kategori, 6=Item Material, 7=Spesifikasi,
# 8=Produk/Merk, 9=Sumber RKS, 10=Prioritas, 11=Status Pengajuan, 12=Status Approval MK,
# 13=Status Approval Owner, 14=Tgl Submit MK, 15=Tgl Approved MK, 16=Tgl Submit Owner,
# 17=Tgl Approved Owner, 18=PIC, 19=No Transmittal, 20=Catatan

print(f"Loaded: {ws.max_row} rows")

# ─── Find rows by ID ───
def find_row_by_id(mat_id):
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value == mat_id:
            return r
    return None

# ─── Updates ───
updates = {
    "AM-GD-STR-104": {  # Tiang pancang
        7: "Beton mutu fc' 45 Mpa; Baja Tulangan Ulir BJTS 420; Mutu strand ASTM A416 grade 270",
        8: "PPI, JHS, Tripalindo P",
    },
    "AM-GD-STR-106": {  # Beton ready mix
        7: "fc' = 30 Mpa (Pile cap, dinding beton, kolom, balok, pelat, tiang bor); fc' = 45 Mpa khusus Tiang Pancang; Semen Portland Tipe 1; Agregat SNI 8321-2016 / ASTM C-33; Air SNI 7974:2016",
        8: "Ready Mix: Karya Beton, Jaya Mix",
    },
    "AM-GD-STR-107": {  # Besi tulangan
        7: "BJTS-420 (Yield Strength min. 420 Mpa, maks. 545 Mpa; Kuat tarik min. 525 Mpa); SNI 2052:2024",
        8: "Master Steel, Interwood Steel, Cakra Tunggal Steel, Jaya Steel, Krakatau Steel",
    },
    "AM-GD-STR-111": {  # Baja struktur
        7: "Pipa Schedule A-36; Baja Profil A-36; Las AWS E-70XX; Baut ASTM A-325",
        8: "Gunung Garuda, Hanin Jaya Steel, Krakatau Osaka Steel, Garuda Yamato Steel",
    },
    "AM-GD-STR-114": {  # Waterproofing struktur
        7: "MEMBRANE: Sika Bitusel T-130SG, Tamseal, Fosroc, Penetron atau setara (plat atap & plat exterior); SPRAY: Sika, Fosroc, Penetron; COATING: Sika, Fosroc, Penetron; INTEGRAL: untuk GWT & STP",
        8: "Sika, Fosroc, Penetron atau setara",
    },
}

updated = 0
for mat_id, changes in updates.items():
    row = find_row_by_id(mat_id)
    if row:
        for col, val in changes.items():
            ws.cell(row, col, val)
        print(f"  Updated: {mat_id} (row {row})")
        updated += 1
    else:
        print(f"  NOT FOUND: {mat_id}")

# ─── Add new materials ───
# Find the last row with data in STR section
last_str_row = None
for r in range(ws.max_row, 1, -1):
    if ws.cell(r, 4).value == "STR" and ws.cell(r, 2).value:
        last_str_row = r
        break

print(f"\nLast STR row: {last_str_row}")

# New materials to add
new_materials = [
    {
        2: "AM-GD-STR-175", 3: "Gedung D", 4: "STR", 5: "Semen",
        6: "Semen Portland",
        7: "Warna abu-abu, bentuk powder; SNI 2049:2015 / ASTM C 150/C150M-12 / BS 197-1:2000; Kemasan 40 kg, 50 kg",
        8: "Semen Tiga Roda, Semen Gresik, Semen Merah Putih, Semen Padang",
        9: "Outline Spek Struktur - A.1",
        10: "Tinggi", 11: "Belum Diajukan", 12: "Belum Approved", 13: "Belum Approved",
        20: "Siapkan submittal: sample/brosur/COA sesuai RKS."
    },
    {
        2: "AM-GD-STR-176", 3: "Gedung D", 4: "STR", 5: "Beton",
        6: "Beton fc' 45 Tiang Pancang",
        7: "Kuat tekan 45 Mpa khusus Tiang Pancang; Semen Portland Tipe 1; Agregat SNI 8321-2016; Air SNI 7974:2016",
        8: "Ready Mix: Karya Beton, Jaya Mix",
        9: "Outline Spek Struktur - A.2",
        10: "Tinggi", 11: "Belum Diajukan", 12: "Belum Approved", 13: "Belum Approved",
        20: "Terpisah dari beton fc'30 struktur umum. Siapkan submittal: mix design, trial mix, slump test, cylinder test."
    },
]

# Add rows after last STR row
for nm in new_materials:
    last_str_row += 1
    # Copy style from row above
    for col in range(1, 21):
        if col in nm:
            ws.cell(last_str_row, col, nm[col])
        else:
            ws.cell(last_str_row, col, None)
    # Set No
    ws.cell(last_str_row, 1, last_str_row - 1)  # sequential no
    print(f"  Added: {nm[2]} (row {last_str_row})")

# ─── Update Summary sheet ───
ws_summary = wb["Summary"]
# Update total item count
total = ws.max_row - 1  # minus header
ws_summary.cell(7, 2, total)
print(f"\nUpdated Summary: Total Item = {total}")

# Update timestamp
ws_summary.cell(3, 2, datetime.now().strftime("%d %B %Y %H:%M"))

# ─── Save ───
wb.save(EXCEL_PATH)
print(f"\n✅ Saved: {EXCEL_PATH}")
print(f"   Updated: {updated} rows")
print(f"   Added: {len(new_materials)} rows")
print(f"   Total: {total} materials")
